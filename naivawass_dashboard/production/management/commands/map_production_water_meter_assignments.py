import os
import re

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from metering.models import ProductionWaterMeterAssignment, WaterMeter
from metering.sync import sync_production_site_metering
from metering.workbook_names import slugify_meter_name, split_meter_name_and_notes
from production.models import ProductionSite, WaterSource


SHEET_SITE_ALIASES = {
    'WATERWORKS': 'WWS',
    'POLICELINE': 'POLICE',
    'KARATEPOLICEPOST': 'KARATIPP',
    'MAAIMAHIU': 'MAIMAHIU',
    'MWICHIRIGIRI': 'MUCHIRINGIRI',
    'MWICHIRIGIRIDISPENSARY': 'MUCHIRIDISP',
    'AIC': 'AIC',
    'IHINDU1': 'IHINDU1',
    'IHINDU2': 'IHINDU2',
    'NGONDI': 'NGONDI',
    'NYONJORO': 'NYONJORO',
    'DTI': 'DTI',
    'DCK': 'DCK',
    'KARATI': 'KARATI',
    'NIP': 'NIP',
    'KIHOTO': 'KIHOTO',
    'MANERA': 'MANERA',
}

SHARED_SUPPLY_MARKERS = (
    'BOTH PRODUCTION AND SUPPLY',
    'BOTH ABSTRACTED AND SUPPLY',
    '0 PRODUCTION LOSS',
)

SITE_MATCH_HINTS = {
    'WWS': ['WWKS '],
    'DTI': ['DTI '],
    'KARATI': ['KARATI BH', 'KARATI -', 'KARATI BULK', 'UPPER KWS '],
    'NIP': ['NIP ', 'MAILI MBILI'],
    'MAIMAHIU': ['MAAI MAHIU '],
    'AIC': ['AIC '],
    'POLICE': ['POLICE LINE '],
    'IHINDU1': ['IHINDU 1 '],
    'IHINDU2': ['IHINDU 2 '],
    'MUCHIRINGIRI': ['MUCHIRINGIRI '],
    'MUCHIRIDISP': ['MUCH-DISP '],
    'NYONJORO': ['NYONJORO '],
    'KARATIPP': ['KARATI PP '],
    'DCK': ['DCK '],
    'NGONDI': ['NGONDI '],
}

LABEL_OVERRIDES = {
    ('KARATI', 'WATERWORKSMETER'): 'KARATI - WWKS METER',
    ('KARATI', 'UPPERKWSWWS'): 'UPPER KWS WWKS METER',
    ('NIP', 'MAILIMBILIMETER'): 'MAILI MBILI SUPPLY METER',
    ('MUCHIRIDISP', 'MUCHIRINGIRIDISPENSARYOUTLETMETER'): 'MUCH-DISP OUTLET METER',
}

SOURCE_CODE_OVERRIDES = {
    ('IHINDU2', 'BH1METER'): 'BH2',
    ('MUCHIRIDISP', 'BH1METER'): 'DISP',
}


def normalize_text(value):
    return re.sub(r'[^A-Z0-9]+', '', str(value or '').upper())


def strip_parenthetical(value):
    return re.sub(r'\s*\([^)]*\)', '', str(value or '')).strip()
class Command(BaseCommand):
    help = (
        'Map shared water meters onto production sites as abstraction and supply assignments '
        'using production sites.xlsx without creating or renaming any water meter records.'
    )

    def add_arguments(self, parser):
        parser.add_argument(
            '--excel-file',
            default=None,
            help='Path to production sites.xlsx. Defaults to ../production sites.xlsx relative to manage.py.',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Parse workbook and report planned assignments without writing to the database.',
        )
        parser.add_argument(
            '--meter-workbook',
            default=None,
            help='Path to WATER METERS UPDATED.xlsx used to restore original imported display names.',
        )

    @transaction.atomic
    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError('openpyxl is required to read the workbook.') from exc

        excel_path = options['excel_file']
        if not excel_path:
            base = os.path.dirname(os.path.abspath('manage.py'))
            excel_path = os.path.normpath(os.path.join(base, '..', 'production sites.xlsx'))

        meter_workbook_path = options['meter_workbook']
        if not meter_workbook_path:
            base = os.path.dirname(os.path.abspath('manage.py'))
            meter_workbook_path = os.path.normpath(os.path.join(base, '..', 'WATER METERS UPDATED.xlsx'))

        if not os.path.exists(excel_path):
            raise CommandError(f'Excel file not found: {excel_path}')
        if not os.path.exists(meter_workbook_path):
            raise CommandError(f'Meter workbook not found: {meter_workbook_path}')

        dry_run = options['dry_run']
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        imported_name_map = self._load_imported_meter_names(meter_workbook_path)
        self.imported_name_map = imported_name_map

        site_lookup = self._build_site_lookup()
        meter_lookup = self._build_meter_lookup()

        assignment_plan = {}
        shared_site_codes = set()
        warnings = []

        for worksheet in workbook.worksheets:
            headers = {
                normalize_text(worksheet.cell(row=1, column=column).value): column - 1
                for column in range(1, worksheet.max_column + 1)
            }
            production_index = headers.get('PRODUCTIONWATERABSTRACTED')
            supply_index = headers.get('WATERSUPPLIED')
            special_index = headers.get('SPECIAL')

            if production_index is None or supply_index is None:
                continue

            site = site_lookup.get(normalize_text(worksheet.title))
            if site is None:
                warnings.append(f'Skipped sheet "{worksheet.title}": no matching production site.')
                continue

            desired = {'ABSTRACTION': [], 'SUPPLY': []}
            if self._sheet_indicates_shared_supply(worksheet, production_index, supply_index, special_index):
                shared_site_codes.add(site.code)

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                production_value = row[production_index] if production_index < len(row) else None
                supply_value = row[supply_index] if supply_index < len(row) else None

                production_meter = self._resolve_meter(site, production_value, meter_lookup)
                if production_value and production_meter is None:
                    warnings.append(
                        f'No imported water meter matched abstraction label "{production_value}" for site {site.code}.'
                    )
                if production_meter:
                    desired['ABSTRACTION'].append({
                        'meter': production_meter,
                        'water_source': self._resolve_water_source(site, str(production_value)),
                        'label': str(production_value).strip(),
                    })

                supply_meter = self._resolve_meter(site, supply_value, meter_lookup)
                if supply_value and supply_meter is None and site.code not in shared_site_codes:
                    warnings.append(
                        f'No imported water meter matched supply label "{supply_value}" for site {site.code}.'
                    )
                if supply_meter:
                    desired['SUPPLY'].append({
                        'meter': supply_meter,
                        'water_source': None,
                        'label': str(supply_value).strip(),
                    })

            assignment_plan[site.code] = desired

        created = 0
        updated = 0
        removed = 0
        flag_updates = 0

        for site_code, desired in assignment_plan.items():
            site = site_lookup[normalize_text(site_code)]
            should_share = site.code in shared_site_codes
            if site.production_equals_supply != should_share:
                site.production_equals_supply = should_share
                flag_updates += 1
                if not dry_run:
                    site.save(update_fields=['production_equals_supply'])

            desired_keys = set()
            for role, rows in desired.items():
                seen = set()
                for row in rows:
                    water_source_id = row['water_source'].id if row['water_source'] else None
                    key = (row['meter'].id, water_source_id, role)
                    if key in seen:
                        continue
                    seen.add(key)
                    desired_keys.add(key)

                    existing = ProductionWaterMeterAssignment.objects.filter(
                        water_meter=row['meter'],
                        production_site=site,
                        water_source=row['water_source'],
                        assignment_role=role,
                    ).first()
                    if existing:
                        changed = False
                        desired_notes = (
                            f'Mapped from production sites.xlsx label "{row["label"]}" '
                            f'as {role.lower()} for {site.code}.'
                        )
                        if not existing.is_active:
                            existing.is_active = True
                            changed = True
                        if existing.notes != desired_notes:
                            existing.notes = desired_notes
                            changed = True
                        if changed:
                            updated += 1
                            if not dry_run:
                                existing.save(update_fields=['is_active', 'notes', 'updated_at'])
                        continue

                    created += 1
                    if not dry_run:
                        ProductionWaterMeterAssignment.objects.create(
                            water_meter=row['meter'],
                            production_site=site,
                            water_source=row['water_source'],
                            assignment_role=role,
                            is_active=True,
                            notes=(
                                f'Mapped from production sites.xlsx label "{row["label"]}" '
                                f'as {role.lower()} for {site.code}.'
                            ),
                        )

            existing_assignments = ProductionWaterMeterAssignment.objects.filter(
                production_site=site
            ).select_related('water_meter', 'water_source')
            for assignment in existing_assignments:
                key = (
                    assignment.water_meter_id,
                    assignment.water_source_id,
                    assignment.assignment_role,
                )
                if key in desired_keys:
                    continue
                removed += 1
                if not dry_run:
                    assignment.delete()

            if not dry_run:
                sync_production_site_metering(site)

        restored_names = 0
        for meter_number, display_name in imported_name_map.items():
            meter = WaterMeter.objects.filter(meter_number=meter_number).first()
            if meter is None or meter.display_name == display_name:
                continue
            restored_names += 1
            if not dry_run:
                meter.display_name = display_name
                meter.save(update_fields=['display_name', 'updated_at'])

        if dry_run:
            transaction.set_rollback(True)

        for warning in warnings:
            self.stdout.write(self.style.WARNING(warning))

        self.stdout.write(self.style.SUCCESS(
            f'Production water meter mapping complete. '
            f'Flags updated: {flag_updates}. Assignments created: {created}. '
            f'Assignments updated: {updated}. Assignments removed: {removed}. '
            f'Display names restored: {restored_names}. '
            f'Shared-meter sites: {", ".join(sorted(shared_site_codes)) if shared_site_codes else "none"}.'
        ))

    def _build_site_lookup(self):
        lookup = {}
        for site in ProductionSite.objects.all():
            lookup[normalize_text(site.code)] = site
            lookup[normalize_text(site.name)] = site

        for alias, site_code in SHEET_SITE_ALIASES.items():
            site = ProductionSite.objects.filter(code=site_code).first()
            if site:
                lookup[alias] = site
        return lookup

    def _build_meter_lookup(self):
        return {
            normalize_text(meter.display_name): meter
            for meter in WaterMeter.objects.order_by('display_name')
        }

    def _load_imported_meter_names(self, workbook_path):
        import openpyxl

        workbook = openpyxl.load_workbook(workbook_path, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        mapping = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            raw_value = row[0] if row else None
            raw_display_name = str(raw_value).strip() if raw_value is not None else ''
            if not raw_display_name:
                continue
            display_name, _ = split_meter_name_and_notes(raw_display_name)
            mapping[slugify_meter_name(raw_display_name)] = display_name
            mapping[slugify_meter_name(display_name)] = display_name
        return mapping

    def _resolve_meter(self, site, raw_label, meter_lookup):
        if not raw_label:
            return None

        label = str(raw_label).strip()
        candidates = []
        search_variants = [
            normalize_text(label),
            normalize_text(strip_parenthetical(label)),
        ]
        search_variants = [variant for variant in search_variants if variant]

        for variant in search_variants:
            override = LABEL_OVERRIDES.get((site.code, variant))
            if override:
                meter = WaterMeter.objects.filter(display_name=override).first()
                if meter:
                    return meter

            exact = meter_lookup.get(variant)
            if exact:
                return exact

        for meter in WaterMeter.objects.order_by('display_name'):
            meter_key = normalize_text(meter.display_name)
            for variant in search_variants:
                if meter_key.endswith(variant) or variant.endswith(meter_key) or variant in meter_key or meter_key in variant:
                    candidates.append(meter)
                    break

        imported_candidates = []
        for meter_number, imported_name in getattr(self, 'imported_name_map', {}).items():
            imported_key = normalize_text(imported_name)
            for variant in search_variants:
                if (
                    imported_key.endswith(variant)
                    or variant.endswith(imported_key)
                    or variant in imported_key
                    or imported_key in variant
                ):
                    imported_candidates.append((meter_number, imported_name))
                    break

        if imported_candidates:
            preferred_imported = []
            for hint in SITE_MATCH_HINTS.get(site.code, []):
                preferred_imported.extend(
                    candidate for candidate in imported_candidates
                    if candidate[1].upper().startswith(hint)
                )
            if preferred_imported:
                imported_candidates = preferred_imported

            deduped = []
            seen_numbers = set()
            for meter_number, imported_name in imported_candidates:
                if meter_number in seen_numbers:
                    continue
                seen_numbers.add(meter_number)
                deduped.append((meter_number, imported_name))

            if len(deduped) == 1:
                meter = WaterMeter.objects.filter(meter_number=deduped[0][0]).first()
                if meter:
                    return meter

            if not candidates and deduped:
                raise CommandError(
                    f'Ambiguous imported meter label "{label}" for site {site.code}. Candidates: '
                    + ', '.join(imported_name for _, imported_name in deduped)
                )

        if not candidates:
            return None

        preferred = []
        for hint in SITE_MATCH_HINTS.get(site.code, []):
            preferred.extend(
                meter for meter in candidates
                if meter.display_name.upper().startswith(hint)
            )
        if preferred:
            candidates = preferred

        unique_candidates = []
        seen_ids = set()
        for meter in candidates:
            if meter.id in seen_ids:
                continue
            seen_ids.add(meter.id)
            unique_candidates.append(meter)

        if len(unique_candidates) == 1:
            return unique_candidates[0]

        raise CommandError(
            f'Ambiguous meter label "{label}" for site {site.code}. Candidates: '
            + ', '.join(meter.display_name for meter in unique_candidates)
        )

    def _resolve_water_source(self, site, raw_label):
        sources = list(site.water_sources.all().order_by('code'))
        if not sources:
            return None

        label = str(raw_label or '').upper()
        override_code = SOURCE_CODE_OVERRIDES.get((site.code, normalize_text(label)))
        if override_code:
            for source in sources:
                if normalize_text(source.code) == normalize_text(override_code):
                    return source

        bh_match = re.search(r'BH\s*([0-9]+)', label)
        if not bh_match:
            borehole_match = re.search(r'BOREHOLE\s*([0-9]+)', label)
            if borehole_match:
                bh_match = borehole_match
        if bh_match:
            source_code = f'BH{bh_match.group(1)}'
            for source in sources:
                if normalize_text(source.code) == normalize_text(source_code):
                    return source
            return None
        if len(sources) == 1:
            return sources[0]
        return None

    def _sheet_indicates_shared_supply(self, worksheet, production_index, supply_index, special_index):
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            special_value = row[special_index] if special_index is not None and special_index < len(row) else None
            special_text = str(special_value or '').upper()

            if any(marker in special_text for marker in SHARED_SUPPLY_MARKERS):
                return True
        return False
