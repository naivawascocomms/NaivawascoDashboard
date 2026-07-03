"""
Import historical production data from the Excel workbook.

Processes three sheets:
  1. 'production'  → DailyProduction.water_abstracted_m3
  2. 'Kwh'         → DailyProduction.power_grid_kwh  (total energy per site)
  3. 'Supply'      → MeterReading (opening/closing per supply meter per day)

Both production and Kwh data are collected in memory first so that each
DailyProduction record is written once via update_or_create (triggering
save() for derived field calculation).

After daily records are complete, aggregates into MonthlyProduction.

Usage:
    python manage.py import_production_excel <path_to_xlsx>

Flags:
    --dry-run     Parse and report counts without writing to DB
    --clear       Delete existing DailyProduction + MeterReading before importing
"""
import re
from datetime import date, datetime, time as dt_time
from decimal import Decimal, InvalidOperation
from collections import defaultdict

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from production.models import (
    ProductionSite, Meter, MeterReading, DailyProduction, MonthlyProduction,
)

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ---------------------------------------------------------------------------
# Site name → ProductionSite.code  (strips whitespace before lookup)
# ---------------------------------------------------------------------------
SITE_NAME_MAP = {
    'DCK':         'DCK',
    'WWS':         'WWS',
    'Mai Mahiu':   'MAI_MAHIU',
    'Ngondi':      'NGONDI',
    'NIP':         'NIP',
    'KARATI':      'KARATI',
    'Kinungi BH':  'KINUNGI_BH',
    'Nyonjoro':    'NYONJORO',
    'Karati PP':   'KARATI_PP',
    'AIC':         'AIC',
    'DTI':         'DTI',
    'Police Line': 'POLICE_LINE',
}

# Rows whose first cell means "skip this row"
SKIP_FIRST_CELL = {
    'Daily production sheet', 'Site/Date', 'Sub-Total', 'TOTAL', 'TOTALS',
    'Southern', 'Eastern', 'Central BL',
    'WWS+Karati (Mirera)', 'WWS+KARATI( Kayole)', 'WWS+Karati (L.V)',
}

# Columns carrying actual daily values (skipping week-total at col 8 and 16)
DAILY_COLS = list(range(1, 8)) + list(range(9, 16))  # [1..7, 9..15]

# Meter name as it appears in the Supply sheet → meter_number in DB
METER_NAME_MAP = {
    'DTI high lift Meter': 'DTI-HIGH-LIFT',
    'AIC':                 'AIC-SUPPLY',
    'Police Line':         'POLICE-LINE-SUPPLY',
    'WWS':                 'WWS-SUPPLY',
    'KCC':                 'KCC-SUPPLY',
    'BULK':                'BULK-SUPPLY',
    'Keroche':             'KEROCHE-SUPPLY',
    'Staff':               'STAFF-SUPPLY',
    'Hopewell':            'HOPEWELL-SUPPLY',
    'Karati Supply':       'KARATI-SUPPLY',
    "WWS 6'' Mirera":      'WWS-6IN-MIRERA',
    "WWS 4'' Sub":         'WWS-4IN-SUB',
    'KWS':                 'KWS-SUPPLY',
    'Office':              'OFFICE-SUPPLY',
    'GEUST INN':           'GUEST-INN-SUPPLY',
    'New line':            'NEW-LINE-SUPPLY',
    'WWS-kayole':          'WWS-KAYOLE',
    'DCK':                 'DCK-SUPPLY',
    'DCK ':                'DCK-SUPPLY',
}


# ---------------------------------------------------------------------------
# Date parsing helpers
# ---------------------------------------------------------------------------

_ORDINAL_RE = re.compile(r'(\d+)(st|nd|rd|th)', re.IGNORECASE)


def _strip_ordinal(s: str) -> str:
    return _ORDINAL_RE.sub(r'\1', s).strip()


def parse_date_no_year(raw: str, tracker: dict):
    """
    Parse '29th December', '1st Jan', '12th Feb', etc.
    tracker = {'year': int, 'last_month': int|None}
    Increments year when month wraps Dec -> Jan.
    """
    if not raw or not isinstance(raw, str):
        return None
    cleaned = _strip_ordinal(raw)
    for fmt in ('%d %B', '%d %b'):
        try:
            dt = datetime.strptime(cleaned, fmt)
            month = dt.month
            last = tracker['last_month']
            if last is not None and month < last and (last - month) > 3:
                tracker['year'] += 1
            tracker['last_month'] = month
            return date(tracker['year'], month, dt.day)
        except ValueError:
            continue
    return None


def parse_date_with_year(raw: str):
    """Parse '28th December 2023', '1st January 2024', etc."""
    if not raw or not isinstance(raw, str):
        return None
    cleaned = _strip_ordinal(raw)
    for fmt in ('%d %B %Y', '%d %b %Y'):
        try:
            return datetime.strptime(cleaned, fmt).date()
        except ValueError:
            continue
    return None


def safe_decimal(val):
    """Return Decimal for numeric values; None for errors/missing."""
    if val is None:
        return None
    if isinstance(val, str):
        return None  # formula errors like '#REF!', '#VALUE!'
    try:
        return Decimal(str(val))
    except InvalidOperation:
        return None


# ---------------------------------------------------------------------------
# Generic sheet parser for 'production' and 'Kwh' (same layout)
# ---------------------------------------------------------------------------

def parse_production_style_sheet(ws, start_year=2023, start_month=12):
    """
    Returns dict: {(site_code, date): value}
    """
    tracker = {'year': start_year, 'last_month': start_month}
    current_dates = []
    result = {}

    for row in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row):
            continue
        first = row[0]
        if first is None:
            continue
        first_str = str(first).strip()

        if first_str == 'Site/Date':
            current_dates = []
            for col_idx in DAILY_COLS:
                val = row[col_idx] if col_idx < len(row) else None
                if val and isinstance(val, str):
                    s = val.strip()
                    if 'week' not in s.lower() and 'total' not in s.lower():
                        d = parse_date_no_year(s, tracker)
                        if d:
                            current_dates.append(d)
            continue

        if first_str in SKIP_FIRST_CELL or 'Sub-Total' in first_str:
            continue

        site_code = SITE_NAME_MAP.get(first_str)
        if not site_code:
            continue

        for i, d in enumerate(current_dates):
            if i >= len(DAILY_COLS):
                break
            col_idx = DAILY_COLS[i]
            val = safe_decimal(row[col_idx] if col_idx < len(row) else None)
            if val is None:
                val = Decimal('0')
            key = (site_code, d)
            result[key] = result.get(key, Decimal('0')) + val

    return result


# ---------------------------------------------------------------------------
# Main command
# ---------------------------------------------------------------------------

class Command(BaseCommand):
    help = 'Import historical production data from the Excel workbook'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', help='Path to the Excel workbook')
        parser.add_argument('--dry-run', action='store_true',
                            help='Parse only, do not write to DB')
        parser.add_argument('--clear', action='store_true',
                            help='Delete all existing DailyProduction and MeterReading records first')

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError('openpyxl is not installed. Run: pip install openpyxl')

        path = options['excel_file']
        dry_run = options['dry_run']
        clear = options['clear']

        self.stdout.write(f'Loading workbook: {path}')
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception as exc:
            raise CommandError(f'Cannot open workbook: {exc}')

        self.site_map = {s.code: s for s in ProductionSite.objects.all()}
        self.meter_map = {m.meter_number: m for m in Meter.objects.filter(meter_type='SUPPLY')}

        if not self.site_map:
            raise CommandError(
                'No ProductionSite records found. '
                'Run first: python manage.py seed_reference_data'
            )

        # ---- Parse all sheets into memory ----
        self.stdout.write('Parsing production sheet...')
        vol_data = parse_production_style_sheet(wb['production'])
        self.stdout.write(f'  {len(vol_data)} site-day volume records parsed')

        self.stdout.write('Parsing Kwh sheet...')
        kwh_data = parse_production_style_sheet(wb['Kwh'])
        self.stdout.write(f'  {len(kwh_data)} site-day energy records parsed')

        self.stdout.write('Parsing Supply sheet...')
        supply_records = self._parse_supply_sheet(wb['Supply'])
        self.stdout.write(f'  {len(supply_records)} meter reading records parsed')

        if dry_run:
            self.stdout.write(self.style.WARNING('\n[DRY RUN] No data written.'))
            self._print_summary(len(vol_data), 0, len(supply_records), 0, 0)
            return

        # ---- Write to DB ----
        with transaction.atomic():
            if clear:
                self.stdout.write(self.style.WARNING('--clear: removing existing records...'))
                DailyProduction.objects.all().delete()
                MeterReading.objects.all().delete()

            supply_totals = self._build_supply_totals(supply_records)

            self.stdout.write('Writing DailyProduction records...')
            dp_created, dp_updated = self._write_daily_production(vol_data, kwh_data, supply_totals)

            self.stdout.write('Writing MeterReading records...')
            mr_created = self._write_meter_readings(supply_records)

            self.stdout.write('Aggregating MonthlyProduction...')
            mp_created, mp_updated = self._aggregate_monthly()

        self._print_summary(dp_created, dp_updated, mr_created, mp_created, mp_updated)

    # ------------------------------------------------------------------

    def _write_daily_production(self, vol_data, kwh_data, supply_totals):
        created = updated = 0
        # Union of all keys from both sheets
        all_keys = set(vol_data) | set(kwh_data) | set(supply_totals)

        for site_code, prod_date in all_keys:
            site = self.site_map.get(site_code)
            if not site:
                continue
            volume = vol_data.get((site_code, prod_date), Decimal('0'))
            kwh    = kwh_data.get((site_code, prod_date), Decimal('0'))
            supplied = supply_totals.get((site_code, prod_date), Decimal('0'))

            obj, was_created = DailyProduction.objects.update_or_create(
                production_site=site,
                production_date=prod_date,
                defaults={
                    'water_abstracted_m3': volume,
                    'water_supplied_m3':   supplied,
                    'power_grid_kwh':      kwh,
                    'is_complete':         True,
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1

        return created, updated

    def _build_supply_totals(self, supply_records):
        totals = defaultdict(lambda: Decimal('0'))

        for meter_number, reading_date, _opening, _closing, supply in supply_records:
            meter = self.meter_map.get(meter_number)
            if not meter or supply is None:
                continue
            key = (meter.production_site.code, reading_date)
            totals[key] += supply

        return totals

    def _parse_supply_sheet(self, ws):
        """
        Returns list of (meter_number, date, opening, closing, supply).
        Each row in the sheet has a group every 5 columns starting at col 1:
          [meter_name, opening, closing, supply, separator]
        """
        GROUP_STRIDE = 5
        GROUP_START  = 1

        current_dates = []
        records = []

        for row in ws.iter_rows(values_only=True):
            if not any(v is not None for v in row):
                continue

            col1 = row[GROUP_START] if len(row) > GROUP_START else None

            # Date header row
            if col1 == 'Date ':
                current_dates = []
                col = GROUP_START
                while col + 1 < len(row):
                    if row[col] == 'Date ':
                        raw = row[col + 1]
                        d = parse_date_with_year(str(raw).strip() if raw else '')
                        if d:
                            current_dates.append(d)
                    col += GROUP_STRIDE
                continue

            # Sub-header row (OPENING / CLOSING / SUPPLY labels)
            if col1 is None and len(row) > GROUP_START + 1 and row[GROUP_START + 1] == 'OPENING':
                continue

            # Skip known non-meter total rows
            if col1 in ('TOTAL SUPPLY', 'WWS SUPPLY', None):
                continue

            if not isinstance(col1, str) or not col1.strip():
                continue

            meter_raw = col1.strip()
            if meter_raw in ('Date ', 'OPENING'):
                continue

            meter_number = METER_NAME_MAP.get(meter_raw) or METER_NAME_MAP.get(meter_raw.strip())
            if not meter_number:
                continue

            for group_idx, d in enumerate(current_dates):
                base = GROUP_START + group_idx * GROUP_STRIDE
                if base + 3 >= len(row):
                    break
                opening = safe_decimal(row[base + 1])
                closing = safe_decimal(row[base + 2])
                supply  = safe_decimal(row[base + 3])

                if closing is None:
                    continue

                records.append((meter_number, d, opening, closing, supply))

        return records

    def _write_meter_readings(self, supply_records):
        created = 0
        for meter_number, reading_date, opening, closing, supply in supply_records:
            meter = self.meter_map.get(meter_number)
            if not meter:
                continue
            _, was_created = MeterReading.objects.get_or_create(
                meter=meter,
                reading_date=reading_date,
                reading_time=dt_time(8, 0),
                defaults={
                    'current_reading':  closing,
                    'previous_reading': opening,
                    'consumption':      supply,
                    'reading_method':   'MANUAL',
                },
            )
            if was_created:
                created += 1
        return created

    def _aggregate_monthly(self):
        created = updated = 0

        buckets = defaultdict(lambda: {
            'water':     Decimal('0'),
            'supplied':  Decimal('0'),
            'received':  Decimal('0'),
            'grid_kwh':  Decimal('0'),
            'solar_kwh': Decimal('0'),
        })

        for dp in DailyProduction.objects.select_related('production_site').iterator():
            key = (dp.production_site_id, dp.production_date.year, dp.production_date.month)
            buckets[key]['water']     += dp.water_abstracted_m3 or Decimal('0')
            buckets[key]['supplied']  += dp.water_supplied_m3 or Decimal('0')
            buckets[key]['received']  += dp.water_received_m3 or Decimal('0')
            buckets[key]['grid_kwh']  += dp.power_grid_kwh      or Decimal('0')
            buckets[key]['solar_kwh'] += dp.power_solar_kwh     or Decimal('0')

        site_id_map = {s.id: s for s in ProductionSite.objects.all()}

        for (site_id, year, month), totals in buckets.items():
            site = site_id_map.get(site_id)
            if not site:
                continue
            obj, was_created = MonthlyProduction.objects.update_or_create(
                production_site=site,
                year=year,
                month=month,
                defaults={
                    'water_abstracted_m3': totals['water'],
                    'water_supplied_m3':   totals['supplied'],
                    'water_received_m3':   totals['received'],
                    'power_grid_kwh':      totals['grid_kwh'],
                    'power_solar_kwh':     totals['solar_kwh'],
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1

        return created, updated

    def _print_summary(self, dp_created, dp_updated, mr_created, mp_created, mp_updated):
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS('=== Import summary ==='))
        self.stdout.write(f'  DailyProduction   created : {dp_created}')
        self.stdout.write(f'  DailyProduction   updated : {dp_updated}')
        self.stdout.write(f'  MeterReading      created : {mr_created}')
        self.stdout.write(f'  MonthlyProduction created : {mp_created}')
        self.stdout.write(f'  MonthlyProduction updated : {mp_updated}')
