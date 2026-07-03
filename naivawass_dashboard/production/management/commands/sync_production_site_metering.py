import os
import re

from django.core.management.base import BaseCommand, CommandError

from metering.sync import sync_production_site_metering
from production.models import ProductionSite


SHEET_SITE_ALIASES = {
    'WATERWORKS': 'WWS',
    'POLICELINE': 'POLICE',
    'KARATEPOLICEPOST': 'KARATIPP',
    'MAAIMAHIU': 'MAIMAHIU',
    'MWICHIRIGIRI': 'MUCHIRINGIRI',
    'MWICHIRIGIRIDISPENSARY': 'MUCHIRIDISP',
    'AIC': 'AIC',
    'IHINDU1': 'IHINDU1',
    'IHINDU2': 'IHINDU2',
    'NGONDI': 'NGONDI',
    'NYONJORO': 'NYONJORO',
    'DTI': 'DTI',
    'DCK': 'DCK',
    'KARATI': 'KARATI',
    'NIP': 'NIP',
    'KIHOTO': 'KIHOTO',
    'MANERA': 'MANERA',
}

SHARED_SUPPLY_MARKERS = (
    'BOTH PRODUCTION AND SUPPLY',
    'BOTH ABSTRACTED AND SUPPLY',
    '0 PRODUCTION LOSS',
)


def normalize_text(value):
    return re.sub(r'[^A-Z0-9]+', '', str(value or '').upper())


class Command(BaseCommand):
    help = (
        'Sync production sites whose abstraction and supply are measured by the same meter '
        'from the production sites workbook.'
    )

    def add_arguments(self, parser):
        parser.add_argument(
            '--excel-file',
            default=None,
            help='Path to production sites.xlsx. Defaults to ../production sites.xlsx relative to manage.py.',
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError('openpyxl is required to read the workbook.') from exc

        excel_path = options['excel_file']
        if not excel_path:
            base = os.path.dirname(os.path.abspath('manage.py'))
            excel_path = os.path.normpath(os.path.join(base, '..', 'production sites.xlsx'))

        if not os.path.exists(excel_path):
            raise CommandError(f'Excel file not found: {excel_path}')

        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        shared_codes = self._extract_shared_supply_site_codes(workbook)

        updated = 0
        for site in ProductionSite.objects.order_by('code'):
            should_share = site.code in shared_codes
            if site.production_equals_supply != should_share:
                site.production_equals_supply = should_share
                site.save(update_fields=['production_equals_supply'])
                updated += 1
            else:
                sync_production_site_metering(site)

        shared_list = ', '.join(sorted(shared_codes)) if shared_codes else 'none'
        self.stdout.write(self.style.SUCCESS(
            f'Synced production_equals_supply for {updated} site(s). Shared-meter sites: {shared_list}'
        ))

    def _extract_shared_supply_site_codes(self, workbook):
        shared_codes = set()
        site_lookup = self._build_site_lookup()

        for worksheet in workbook.worksheets:
            if worksheet.max_row < 2:
                continue

            header_map = {
                normalize_text(worksheet.cell(row=1, column=column).value): column - 1
                for column in range(1, worksheet.max_column + 1)
            }
            production_index = header_map.get('PRODUCTIONWATERABSTRACTED')
            supply_index = header_map.get('WATERSUPPLIED')
            special_index = header_map.get('SPECIAL')

            if production_index is None or supply_index is None:
                continue

            sheet_key = normalize_text(worksheet.title)
            site = site_lookup.get(sheet_key)
            if site is None:
                self.stdout.write(self.style.WARNING(
                    f'Workbook sheet "{worksheet.title}" could not be matched to a production site.'
                ))
                continue

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                production_value = row[production_index] if production_index < len(row) else None
                supply_value = row[supply_index] if supply_index < len(row) else None
                special_value = row[special_index] if special_index is not None and special_index < len(row) else None

                if self._row_indicates_shared_supply(production_value, supply_value, special_value):
                    shared_codes.add(site.code)
                    break

        return shared_codes

    def _build_site_lookup(self):
        lookup = {}
        for site in ProductionSite.objects.all():
            lookup[normalize_text(site.code)] = site
            lookup[normalize_text(site.name)] = site

        for alias, site_code in SHEET_SITE_ALIASES.items():
            site = ProductionSite.objects.filter(code=site_code).first()
            if site:
                lookup[alias] = site

        return lookup

    def _row_indicates_shared_supply(self, production_value, supply_value, special_value):
        production_key = normalize_text(production_value)
        supply_key = normalize_text(supply_value)
        special_text = str(special_value or '').upper()

        if production_key and supply_key and production_key == supply_key:
            return True

        return any(marker in special_text for marker in SHARED_SUPPLY_MARKERS)
