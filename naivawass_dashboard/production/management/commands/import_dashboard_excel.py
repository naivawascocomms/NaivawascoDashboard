"""
Import production dashboard data from the utility's master Excel workbook.

Handles:
  - Reference data cleanup (regions, duplicate sites)
  - Budget/target import (Budget2526 sheet)
  - Actual production import (all Actual* sheets for FY 2020/21 – 2025/26)
  - Company-level cost and water quality import (CompanyMonthlySummary)

Usage:
  python manage.py import_dashboard_excel "../production dashboard.xlsx"
  python manage.py import_dashboard_excel "../production dashboard.xlsx" --dry-run
"""

import re
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand
from django.db import transaction

import openpyxl

from production.models import (
    CompanyMonthlySummary,
    DailyProduction,
    MonthlyProduction,
    ProductionSite,
    ProductionTarget,
    Region,
)


# ────────────────────────────────────────────────────────────────────
# Site definitions: name → (code, region_code, has_solar, has_grid)
# Regional groupings derived from numerical analysis of the Excel
# budget vs actual regional "available for sale" figures.
# ────────────────────────────────────────────────────────────────────
SITE_DEFS = {
    "DTI":                 ("DTI",       "CENTRAL", True,  True),
    "A.I.C":               ("AIC",       "CENTRAL", False, True),
    "KARATI":              ("KARATI",    "CENTRAL", True,  True),
    "NIP":                 ("NIP",       "SOUTH",   False, True),
    "DCK":                 ("DCK",       "SOUTH",   True,  True),
    "MAI MAHIU":           ("MAIMAHIU",  "SOUTH",   False, True),
    "NGONDI":              ("NGONDI",    "SOUTH",   True,  False),
    "MUCHIRINGIRI":        ("MUCHIRI",   "SOUTH",   False, False),
    "MANERA":              ("MANERA",    "SOUTH",   True,  True),
    "KIHOTO":              ("KIHOTO",    "SOUTH",   True,  True),
    "WATERWORKS":          ("WWS",       "EAST",    True,  True),
    "POLICELINE":          ("POLICE",    "EAST",    True,  True),
    "KARATI POLICE POST":  ("KARATIPP",  "EAST",    False, True),
    "NYONJORO":            ("NYONJORO",  "EAST",    True,  False),
    "IHINDU":              ("IHINDU",    "EAST",    False, True),
}

# Excel site header text → canonical site name
HEADER_ALIASES = {
    "DTI PRODUCTION SITE":          "DTI",
    "DTI":                          "DTI",
    "WATERWORKS PRODUCTION SITE":   "WATERWORKS",
    "WATERWORKS":                   "WATERWORKS",
    "A.I.C PRODUCTION SITE":        "A.I.C",
    "A.I.C":                        "A.I.C",
    "AIC PRODUCTION SITE":          "A.I.C",
    "KARATI PRODUCTION SITE":       "KARATI",
    "KARATI":                       "KARATI",
    "DCK PRODUCTION SITE":          "DCK",
    "DCK":                          "DCK",
    "POLICELINE PRODUCTION SITE":   "POLICELINE",
    "POLICELINE":                   "POLICELINE",
    "MAI MAHIU PRODUCTION SITE":    "MAI MAHIU",
    "MAI MAHIU":                    "MAI MAHIU",
    "MUCHIRINGIRI PRODUCTION SITE": "MUCHIRINGIRI",
    "MUCHIRINGIRI":                 "MUCHIRINGIRI",
    "NGONDI":                       "NGONDI",
    "NYONJORO":                     "NYONJORO",
    "KARATI POLICE POST":           "KARATI POLICE POST",
    "NIP":                          "NIP",
    "IHINDU":                       "IHINDU",
    "MANERA PRODUCTION SITE":       "MANERA",
    "MANERA":                       "MANERA",
    "KIHOTO PRODUCTION SITE":       "KIHOTO",
    "KIHOTO":                       "KIHOTO",
}

# ── Label → field mapping for per-site KPI rows ────────────────────
# We only import raw input fields; derived fields are computed in save().
LABEL_MAP = {
    "Water Volume Abstracted":   "water_abstracted_m3",
    "Water Volume Supplied":     "water_supplied",
    "Water Volume Suplied":      "water_supplied",
    "Production Loss":           "production_loss_m3",
    "Power Grid":                "power_grid_kwh",
    "Power Consumption":         "power_grid_kwh",   # non-solar sites
    "Power Solar":               "power_solar_kwh",
    "Power solar":               "power_solar_kwh",
    "Power solar ":              "power_solar_kwh",
    "Power Solar ":              "power_solar_kwh",
}

# ── Company-level label → field mapping (sections 19–20) ──────────
COMPANY_COST_LABELS = {
    "Power costs":                             "power_costs",
    "Repair & Maintenance Production":         "repair_maintenance_costs",
    "Abstraction fee":                         "abstraction_fee",
    "Chemicals costs":                         "chemical_costs",
    "Total direct costs WATER":                "total_direct_costs",
    "Total direct costs WATER per m3 water produced": "total_cost_per_m3",
    "Power costs per m3 water produced":       "power_cost_per_m3",
    "Power costs per kwh":                     "power_cost_per_kwh",
}

# FY label → (start_year, sheet_type)
# FY 2025/26 starts Jul 2025
ACTUAL_SHEETS = {
    "ACTUAL 2526": 2025,
    "Actual2425":  2024,
    "ACTUAL 2324": 2023,
    "Actual2223":  2022,
    "Actual2122":  2021,
    "Actual2021":  2020,
}

MONTH_NAMES = ["jul", "aug", "sep", "oct", "nov", "dec",
               "jan", "feb", "mar", "apr", "may", "jun"]


def _d(val):
    """Convert a value to Decimal, returning 0 for None/errors."""
    if val is None:
        return Decimal("0")
    try:
        s = str(val).strip()
        if s in ("", "#DIV/0!", "#REF!", "#N/A", "#VALUE!"):
            return Decimal("0")
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _fy_year_month(start_year, month_index):
    """Convert FY start year + month index (0=jul … 11=jun) to (year, month)."""
    if month_index < 6:
        return start_year, month_index + 7        # jul=7 … dec=12
    return start_year + 1, month_index - 5         # jan=1 … jun=6


def _identify_site(label):
    """Try to match a cell label to a canonical site name."""
    if not label:
        return None
    s = str(label).strip().upper()
    # Try exact match first
    for alias, canonical in HEADER_ALIASES.items():
        if s == alias.upper():
            return canonical
    # Fuzzy: check if label contains a known site name
    for alias, canonical in HEADER_ALIASES.items():
        if alias.upper() in s and "PRODUCTION" not in alias.upper():
            continue  # skip short names for fuzzy
        if alias.upper() in s:
            return canonical
    return None


def _parse_sheet_sites(ws, data_col_start):
    """
    Parse a worksheet and return per-site monthly data.

    Returns: {canonical_site_name: {field: {month_index: Decimal}}}
    data_col_start: column index (0-based) where 'jul' data begins.
    """
    sites_data = {}
    current_site = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        b_val = row[1].value if len(row) > 1 else None
        if not b_val or b_val == 0:
            continue
        b_str = str(b_val).strip()

        # Check if this is a site header row
        site = _identify_site(b_str)
        if site and site in SITE_DEFS:
            # Distinguish header rows from KPI rows: headers usually have
            # "PRODUCTION SITE" or the label IS the site name (no unit in col C)
            c_val = row[2].value if len(row) > 2 else None
            is_header = (
                "PRODUCTION SITE" in b_str.upper()
                or b_str.upper() in [s.upper() for s in SITE_DEFS]
                or (c_val is None or c_val == 0)
            )
            if is_header:
                current_site = site
                if current_site not in sites_data:
                    sites_data[current_site] = {}
                continue

        if current_site is None:
            continue

        # Check if this is a KPI row we care about
        field = None
        for lbl, fld in LABEL_MAP.items():
            if b_str == lbl or b_str.strip() == lbl.strip():
                field = fld
                break

        if field is None:
            # Check for section headers that reset current_site
            upper = b_str.upper()
            if any(kw in upper for kw in [
                "TOTAL ABSTRACTION", "SUMMARY", "DASHBOARD -",
                "BULK WATER", "CLOSING DATE"
            ]):
                current_site = None
            continue

        # Read monthly values
        monthly = {}
        for mi in range(12):
            col_idx = data_col_start + mi
            if col_idx < len(row):
                val = row[col_idx].value
                monthly[mi] = _d(val)

        sites_data[current_site][field] = monthly

    return sites_data


def _parse_company_data(ws, data_col_start):
    """
    Parse company-level cost and quality data from sections 18–20.

    Returns: {field_name: {month_index: Decimal}}
    """
    data = {}
    in_cost_section = False
    in_quality_section = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        b_val = row[1].value if len(row) > 1 else None
        if not b_val or b_val == 0:
            continue
        b_str = str(b_val).strip()
        upper = b_str.upper()

        # Detect section boundaries
        if "PRODUCTION COSTS" in upper or "POWER CONSUMPTION" in upper:
            in_cost_section = True
            in_quality_section = False
            continue
        if "WATER QUALITY" in upper:
            in_quality_section = True
            in_cost_section = False
            continue

        monthly = {}
        for mi in range(12):
            col_idx = data_col_start + mi
            if col_idx < len(row):
                monthly[mi] = _d(row[col_idx].value)

        if in_cost_section:
            for lbl, fld in COMPANY_COST_LABELS.items():
                if b_str.startswith(lbl) or b_str == lbl:
                    data[fld] = monthly
                    break

        if in_quality_section:
            # Quality rows — match by keyword patterns
            lo = b_str.lower()
            if "compliance" not in lo and "tests" not in lo:
                continue
            if "compliance" in lo:
                if "production" in lo and "chemical" in lo:
                    data["who_compliance_chemical_production"] = monthly
                elif "production" in lo and "biological" in lo:
                    data["who_compliance_biological_production"] = monthly
                elif "consumer" in lo and "chemical" in lo:
                    data["who_compliance_chemical_consumer"] = monthly
                elif "consumer" in lo and "biological" in lo:
                    data["who_compliance_biological_consumer"] = monthly
            # Test counts come in pairs: first 4 are budget (20a-20d), next 4 actual (20e-20h)
            # Distinguish by KPI number in col A
            a_val = str(row[0].value).strip() if row[0].value else ""
            if "tests" in lo:
                if "production" in lo and "chemical" in lo:
                    key = "target_chemical_tests_production" if a_val in ("20a",) else "chemical_tests_production"
                    data[key] = monthly
                elif "production" in lo and "biological" in lo:
                    key = "target_biological_tests_production" if a_val in ("20b",) else "biological_tests_production"
                    data[key] = monthly
                elif "consumer" in lo and "chemical" in lo:
                    key = "target_chemical_tests_consumer" if a_val in ("20c",) else "chemical_tests_consumer"
                    data[key] = monthly
                elif "consumer" in lo and "biological" in lo:
                    key = "target_biological_tests_consumer" if a_val in ("20d",) else "biological_tests_consumer"
                    data[key] = monthly

    return data


def _parse_regional_data(ws, data_col_start):
    """
    Parse section 17 regional closing dates and available-for-sale values.

    Returns dict keyed by region prefix (central/southern/eastern).
    """
    regions = {}
    current_region = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        b_val = row[1].value if len(row) > 1 else None
        if not b_val or b_val == 0:
            continue
        b_str = str(b_val).strip()
        upper = b_str.upper()

        if "CLOSING DATE CENTRAL" in upper:
            current_region = "central"
        elif "CLOSING DATE SOUTHERN" in upper:
            current_region = "southern"
        elif "CLOSING DATE EASTERN" in upper:
            current_region = "eastern"
        elif "TOTAL ABSTRACTION" in upper or "DASHBOARD" in upper:
            current_region = None
            continue

        if current_region is None:
            continue

        if "PRODUCTION LOSS" in upper and current_region:
            monthly = {}
            for mi in range(12):
                col_idx = data_col_start + mi
                if col_idx < len(row):
                    monthly[mi] = _d(row[col_idx].value)
            regions.setdefault(current_region, {})
            regions[current_region]["production_loss_m3"] = monthly

        if "AVAILABLE FOR SALE" in upper and current_region:
            monthly = {}
            for mi in range(12):
                col_idx = data_col_start + mi
                if col_idx < len(row):
                    monthly[mi] = _d(row[col_idx].value)
            regions.setdefault(current_region, {})
            regions[current_region]["available_for_sale_m3"] = monthly

    return regions


class Command(BaseCommand):
    help = "Import production dashboard data from the utility Excel workbook."

    def add_arguments(self, parser):
        parser.add_argument("excel_file", type=str)
        parser.add_argument("--dry-run", action="store_true",
                            help="Parse and report without writing to DB")

    def handle(self, *args, **options):
        path = options["excel_file"]
        dry = options["dry_run"]

        self.stdout.write(f"Loading workbook: {path}")
        wb = openpyxl.load_workbook(path, data_only=True)
        self.stdout.write(f"Sheets: {wb.sheetnames}")

        if dry:
            self._dry_run(wb)
            return

        with transaction.atomic():
            self._cleanup_reference_data()
            site_map = self._ensure_sites()
            self._clear_production_data()
            self._import_budget(wb, site_map)
            self._import_all_actuals(wb, site_map)

        self.stdout.write(self.style.SUCCESS("Import complete."))

    # ── Reference data cleanup ───────────────────────────────────────

    def _cleanup_reference_data(self):
        self.stdout.write("Cleaning up reference data...")

        # Delete sites that are not in the Excel dashboard
        for name in ["Gathima", "Kinungi BH", "Kinungi Production Site"]:
            deleted = ProductionSite.objects.filter(name=name).delete()
            if deleted[0]:
                self.stdout.write(f"  Deleted site: {name} ({deleted})")

        # Delete old duplicate "Naivasha Region" sites
        old_sites = ProductionSite.objects.filter(
            name__endswith="Production Site"
        )
        for site in old_sites:
            self.stdout.write(f"  Deleting old duplicate: {site.name}")
            site.delete()

        # Ensure 3 correct regions, remove stale ones
        Region.objects.update_or_create(
            code="CENTRAL",
            defaults={"name": "Central", "description": "Central production region", "is_active": True}
        )
        Region.objects.update_or_create(
            code="SOUTH",
            defaults={"name": "Southern", "description": "Southern production region", "is_active": True}
        )
        Region.objects.update_or_create(
            code="EAST",
            defaults={"name": "Eastern", "description": "Eastern production region", "is_active": True}
        )

        # Delete empty/stale regions (only if no sites reference them)
        for code in ["NVS", "CR"]:
            try:
                r = Region.objects.get(code=code)
                if not r.production_sites.exists():
                    r.delete()
                    self.stdout.write(f"  Deleted stale region: {r.name} ({code})")
            except Region.DoesNotExist:
                pass

    def _ensure_sites(self):
        """Create / update all 15 production sites. Returns {canonical_name: ProductionSite}."""
        site_map = {}
        for canonical, (code, region_code, has_solar, has_grid) in SITE_DEFS.items():
            region = Region.objects.get(code=region_code)

            # Determine site_type
            if has_grid and has_solar:
                site_type = "MIXED"
            elif has_grid:
                site_type = "BOREHOLE"
            elif has_solar:
                site_type = "BOREHOLE"
            else:
                site_type = "SURFACE"  # gravity-fed like Muchiringiri

            site, created = ProductionSite.objects.update_or_create(
                code=code,
                defaults={
                    "name": canonical,
                    "region": region,
                    "has_solar": has_solar,
                    "site_type": site_type,
                    "is_active": True,
                },
            )
            site_map[canonical] = site
            if created:
                self.stdout.write(f"  Created site: {canonical} ({code}) -> {region.name}")
            else:
                self.stdout.write(f"  Updated site: {canonical} ({code}) -> {region.name}")

        return site_map

    def _clear_production_data(self):
        """Clear all existing production records so we import fresh."""
        counts = {
            "MonthlyProduction": MonthlyProduction.objects.all().delete()[0],
            "ProductionTarget": ProductionTarget.objects.all().delete()[0],
            "DailyProduction": DailyProduction.objects.all().delete()[0],
            "CompanyMonthlySummary": CompanyMonthlySummary.objects.all().delete()[0],
        }
        for model, n in counts.items():
            self.stdout.write(f"  Cleared {n} {model} records")

    # ── Budget import ────────────────────────────────────────────────

    def _import_budget(self, wb, site_map):
        if "Budget2526" not in wb.sheetnames:
            self.stdout.write(self.style.WARNING("Budget2526 sheet not found, skipping."))
            return

        self.stdout.write("Importing Budget2526 as ProductionTargets...")
        ws = wb["Budget2526"]
        # Budget columns: E(4)=jul … P(15)=jun
        sites_data = _parse_sheet_sites(ws, data_col_start=4)

        count = 0
        start_year = 2025  # FY 2025/26
        for canonical, fields in sites_data.items():
            site = site_map.get(canonical)
            if not site:
                self.stdout.write(self.style.WARNING(f"  Unknown site in budget: {canonical}"))
                continue

            for mi in range(12):
                year, month = _fy_year_month(start_year, mi)
                abstracted = fields.get("water_abstracted_m3", {}).get(mi, Decimal("0"))
                if abstracted == 0:
                    continue  # Skip months with no target

                loss = fields.get("production_loss_m3", {}).get(mi, Decimal("0"))
                grid = fields.get("power_grid_kwh", {}).get(mi, Decimal("0"))
                solar = fields.get("power_solar_kwh", {}).get(mi, Decimal("0"))

                ProductionTarget.objects.update_or_create(
                    production_site=site, year=year, month=month,
                    defaults={
                        "water_abstraction_target_m3": abstracted,
                        "production_loss_target_m3": loss,
                        "power_grid_target_kwh": grid,
                        "power_solar_target_kwh": solar,
                    },
                )
                count += 1

        self.stdout.write(f"  Created {count} ProductionTarget records")

        # Import company-level budget costs
        self._import_company_budget(wb, start_year)

    def _import_company_budget(self, wb, start_year):
        """Import budget cost and quality targets into CompanyMonthlySummary."""
        ws = wb["Budget2526"]
        cost_data = _parse_company_data(ws, data_col_start=4)

        for mi in range(12):
            year, month = _fy_year_month(start_year, mi)
            defaults = {}
            for field, monthly in cost_data.items():
                val = monthly.get(mi, Decimal("0"))
                if field.startswith("target_") or field.startswith("who_"):
                    defaults[field] = val
                else:
                    # Cost fields from budget go into target_ prefixed fields
                    defaults[f"target_{field}"] = val

            if defaults:
                summary, _ = CompanyMonthlySummary.objects.update_or_create(
                    year=year, month=month, defaults=defaults,
                )

    # ── Actuals import ───────────────────────────────────────────────

    def _import_all_actuals(self, wb, site_map):
        for sheet_name, start_year in ACTUAL_SHEETS.items():
            if sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(f"Sheet {sheet_name} not found, skipping."))
                continue
            self._import_actuals(wb, sheet_name, start_year, site_map)

    def _import_actuals(self, wb, sheet_name, start_year, site_map):
        self.stdout.write(f"Importing {sheet_name} (FY {start_year}/{start_year+1-2000})...")
        ws = wb[sheet_name]
        # Actual columns: D(3)=jul … O(14)=jun
        sites_data = _parse_sheet_sites(ws, data_col_start=3)

        count = 0
        for canonical, fields in sites_data.items():
            site = site_map.get(canonical)
            if not site:
                self.stdout.write(self.style.WARNING(f"  Unknown site: {canonical}"))
                continue

            for mi in range(12):
                year, month = _fy_year_month(start_year, mi)
                abstracted = fields.get("water_abstracted_m3", {}).get(mi, Decimal("0"))
                if abstracted == 0:
                    continue  # No data for this month

                loss = fields.get("production_loss_m3", {}).get(mi, Decimal("0"))
                # If we have supplied but no explicit loss, derive loss
                supplied = fields.get("water_supplied", {}).get(mi)
                if loss == 0 and supplied is not None and supplied > 0:
                    loss = abstracted - supplied
                    if loss < 0:
                        loss = Decimal("0")

                grid = fields.get("power_grid_kwh", {}).get(mi, Decimal("0"))
                solar = fields.get("power_solar_kwh", {}).get(mi, Decimal("0"))

                # Link to target if it exists
                target = ProductionTarget.objects.filter(
                    production_site=site, year=year, month=month
                ).first()

                MonthlyProduction.objects.update_or_create(
                    production_site=site, year=year, month=month,
                    defaults={
                        "water_abstracted_m3": abstracted,
                        "water_supplied_m3": supplied or Decimal("0"),
                        "production_loss_m3": loss,
                        "power_grid_kwh": grid,
                        "power_solar_kwh": solar,
                        "target": target,
                    },
                )
                count += 1

        self.stdout.write(f"  Created {count} MonthlyProduction records")

        # Import company-level actual data
        self._import_company_actuals(wb, sheet_name, start_year)

    def _import_company_actuals(self, wb, sheet_name, start_year):
        """Import company-level actual costs and quality into CompanyMonthlySummary."""
        ws = wb[sheet_name]
        cost_data = _parse_company_data(ws, data_col_start=3)
        regional_data = _parse_regional_data(ws, data_col_start=3)

        for mi in range(12):
            year, month = _fy_year_month(start_year, mi)
            defaults = {}

            # Cost actuals
            for field, monthly in cost_data.items():
                val = monthly.get(mi, Decimal("0"))
                if not field.startswith("target_"):
                    defaults[field] = val
                else:
                    # Quality test targets from the actual sheet
                    defaults[field] = int(val) if val else 0

            # Regional data
            for region_prefix, rdata in regional_data.items():
                for key, monthly in rdata.items():
                    val = monthly.get(mi, Decimal("0"))
                    defaults[f"{region_prefix}_{key}"] = val

            # Only write if we have actual data
            has_data = any(v for v in defaults.values()
                          if v and v != 0 and v != Decimal("0"))
            if has_data:
                CompanyMonthlySummary.objects.update_or_create(
                    year=year, month=month, defaults=defaults,
                )

    # ── Dry run ──────────────────────────────────────────────────────

    def _dry_run(self, wb):
        self.stdout.write(self.style.WARNING("=== DRY RUN ==="))

        if "Budget2526" in wb.sheetnames:
            ws = wb["Budget2526"]
            sites_data = _parse_sheet_sites(ws, data_col_start=4)
            self.stdout.write(f"\nBudget2526: found {len(sites_data)} sites")
            for site, fields in sites_data.items():
                months_with_data = set()
                for fld, monthly in fields.items():
                    for mi, v in monthly.items():
                        if v > 0:
                            months_with_data.add(mi)
                self.stdout.write(
                    f"  {site}: {len(fields)} fields, "
                    f"{len(months_with_data)} months with data"
                )

        for sheet_name, start_year in ACTUAL_SHEETS.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            sites_data = _parse_sheet_sites(ws, data_col_start=3)
            self.stdout.write(
                f"\n{sheet_name} (FY {start_year}/{start_year+1-2000}): "
                f"found {len(sites_data)} sites"
            )
            for site, fields in sites_data.items():
                months_with_data = set()
                for fld, monthly in fields.items():
                    for mi, v in monthly.items():
                        if v > 0:
                            months_with_data.add(mi)
                self.stdout.write(
                    f"  {site}: {len(fields)} fields, "
                    f"{len(months_with_data)} months with data"
                )
