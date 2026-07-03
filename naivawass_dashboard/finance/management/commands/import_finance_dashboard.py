from decimal import Decimal, InvalidOperation
from pathlib import Path
import re

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from finance.models import FinanceMetric, FinanceMonthlyValue, FinanceReport, FinanceSection

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


MONTHS = {
    'JUL': (7, 1),
    'AUG': (8, 2),
    'SEP': (9, 3),
    'OCT': (10, 4),
    'NOV': (11, 5),
    'DEC': (12, 6),
    'JAN': (1, 7),
    'FEB': (2, 8),
    'MAR': (3, 9),
    'APR': (4, 10),
    'MAY': (5, 11),
    'JUN': (6, 12),
}

LABEL_ALIASES = {
    'total billed (ksh)': [
        'total water/sewerage and related sales (ksh)',
    ],
    'total collection for the month': [
        'total collection, water and sewer revenues and other sales',
    ],
    'bulk water': [
        'bulk',
        'bulk water',
    ],
    'sanitation': [
        'meter rent',
    ],
    'new connection fee sewer': [
        'new connection-sewer',
    ],
    'reconnection fee (water+sewer)': [
        'reconnection fee-water',
    ],
    'misc income-printing fee,ca,meter replacement fee,meter testing': [
        'other sales ca,meter testing,printing',
    ],
    'sewer unblocking fee': [
        'unblockage fee',
    ],
    'penalties (water+sewer)': [
        'penalties-water',
    ],
    'amount billed water-central': [
        'amount billed water-dti',
    ],
    'amount billed sewer-central': [
        'amount billed sewer-dti',
    ],
    'amount billed water & sewer-othersales-central': [
        'amount billed water & sewer-othersales-dti',
    ],
    'amount collected-central': [
        'amount collected-dti',
    ],
    'collection efficiency-central': [
        'collection efficiency-dti',
    ],
    'amount billed water-southern': [
        'amount billed water-karate',
    ],
    'amount billed sewer-southern': [
        'amount billed sewer-karate',
    ],
    'amount billed water & sewer-other sales-southern': [
        'amount billed water & sewer-other sales-karate',
    ],
    'amount billed water & sewer-othersales-southern': [
        'amount billed water & sewer-other sales-karate',
    ],
    'amount collected-southern': [
        'amount collected-karate',
    ],
    'collection efficiency-southern': [
        'money in the bank',
    ],
}


def as_decimal(value):
    if value in [None, '']:
        return None, ''
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None, ''
        try:
            return Decimal(stripped.replace(',', '')), ''
        except InvalidOperation:
            return None, stripped
    try:
        return Decimal(str(value)), ''
    except (InvalidOperation, TypeError, ValueError):
        return None, str(value)


def clean_label(value):
    if value is None:
        return ''
    return str(value).strip()


def normalize_label(value):
    return ' '.join(clean_label(value).lower().split())


def metric_kind(unit, label):
    unit_lower = (unit or '').lower()
    label_lower = label.lower()
    if unit_lower in {'%', 'percent', 'percentage'} or 'efficiency' in label_lower:
        return 'PERCENTAGE'
    if unit_lower in {'no', 'number', 'count'} or 'accounts disconnected' in label_lower:
        return 'COUNT'
    if unit_lower == 'date' or 'region' in label_lower and 'amount' not in label_lower:
        return 'DATE'
    if 'ksh' in unit_lower or 'amount' in label_lower or 'sales' in label_lower or 'collection' in label_lower:
        return 'MONEY'
    return 'TEXT'


def scope_for(section_title, label):
    text = f'{section_title} {label}'.lower()
    for region in ['central', 'southern', 'eastern']:
        if region in text:
            return 'REGION', region.title()
    return 'GLOBAL', ''


def cumulative_behavior(kind, label):
    label_lower = label.lower()
    if kind in {'DATE', 'TEXT'} or 'receivables' in label_lower:
        return 'LAST_VALUE'
    if kind == 'PERCENTAGE':
        return 'AVERAGE'
    return 'SUM'


def is_section_row(row_number, code, label, unit):
    if row_number in {5, 24, 30, 31, 38, 45, 53}:
        return True
    if label == '' and unit == '':
        return True
    if code and label == '' and unit == '':
        return True
    return False


def month_columns(ws):
    columns = {}
    for column in range(1, ws.max_column + 1):
        value = ws.cell(4, column).value
        if isinstance(value, str) and value.strip().upper() in MONTHS:
            columns[value.strip().upper()] = column
    return columns


def metric_row_index(ws):
    index = {}
    for row in range(1, ws.max_row + 1):
        label = normalize_label(ws.cell(row, 2).value)
        if not label:
            continue
        index.setdefault(label, []).append(row)
    return index


def infer_fy_start(sheet_name):
    digits = ''.join(re.findall(r'\d+', sheet_name))
    if len(digits) >= 6:
        return int(digits[:4])
    if len(digits) == 4:
        first = int(digits[:2])
        second = int(digits[2:])
        if first == 20 and second > 20:
            return 2000 + second - 1
        return 2000 + first
    return None


def actual_sheets(workbook):
    sheets = []
    for sheet_name in workbook.sheetnames:
        if 'ACTUAL' not in sheet_name.upper():
            continue
        fy_start = infer_fy_start(sheet_name)
        if fy_start:
            sheets.append((fy_start, sheet_name, workbook[sheet_name]))
    return sorted(sheets, key=lambda item: item[0], reverse=True)


def pick_source_row(source_rows, source_row_usage, label):
    normalized_label = normalize_label(label)
    candidate_labels = [normalized_label, *LABEL_ALIASES.get(normalized_label, [])]
    for candidate in candidate_labels:
        available_rows = source_rows.get(candidate, [])
        occurrence = source_row_usage.get(candidate, 0)
        if occurrence < len(available_rows):
            return available_rows[occurrence], candidate
    return None, normalized_label


def sheet_value(ws, row, columns, month_label):
    if not ws or row is None or month_label not in columns:
        return None
    return ws.cell(row, columns[month_label]).value


class Command(BaseCommand):
    help = 'Import the finance dashboard workbook into generic finance KPI tables.'

    def add_arguments(self, parser):
        parser.add_argument('workbook', type=str)
        parser.add_argument('--fy-start', type=int, default=2025)
        parser.add_argument('--name', type=str, default='Finance Dashboard')
        parser.add_argument('--clear', action='store_true')

    @transaction.atomic
    def handle(self, *args, **options):
        if load_workbook is None:
            raise CommandError('openpyxl is required to import finance workbooks.')

        workbook_path = Path(options['workbook'])
        if not workbook_path.exists():
            raise CommandError(f'Workbook not found: {workbook_path}')

        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        required_sheets = ['Dashboard', 'BUDGET2526', 'General']
        missing = [sheet for sheet in required_sheets if sheet not in workbook.sheetnames]
        if missing:
            raise CommandError(f'Missing required sheets: {", ".join(missing)}')

        dashboard = workbook['Dashboard']
        budget = workbook['BUDGET2526']
        general = workbook['General']
        available_actual_sheets = actual_sheets(workbook)
        if not available_actual_sheets:
            raise CommandError('No actual finance sheets found in workbook.')

        budget_columns = month_columns(budget)
        budget_rows = metric_row_index(budget)
        imported_reports = 0
        imported_metrics = 0
        imported_values = 0

        actual_by_fy = {fy_start: ws for fy_start, _, ws in available_actual_sheets}

        for fy_start, sheet_name, actual in available_actual_sheets:
            previous = actual_by_fy.get(fy_start - 1)
            is_current = fy_start == options['fy_start']
            report, _ = FinanceReport.objects.update_or_create(
                name=options['name'],
                fiscal_year_start=fy_start,
                defaults={
                    'current_snapshot_date': dashboard['B2'].value if is_current and hasattr(dashboard['B2'].value, 'year') else None,
                    'current_fiscal_month_index': int(general['B3'].value or 1) if is_current else 12,
                    'source_workbook': f'{workbook_path.name}:{sheet_name}',
                    'is_active': True,
                },
            )

            if options['clear']:
                report.metrics.all().delete()
                report.sections.all().delete()

            actual_columns = month_columns(actual)
            previous_columns = month_columns(previous) if previous else {}
            source_rows = {
                'budget': budget_rows,
                'actual': metric_row_index(actual),
                'previous': metric_row_index(previous) if previous else {},
            }
            source_row_usage = {
                'budget': {},
                'actual': {},
                'previous': {},
            }

            current_section = None
            section_order = 0
            metric_count = 0
            value_count = 0

            for row in range(5, dashboard.max_row + 1):
                code_raw = dashboard.cell(row, 1).value
                label = clean_label(dashboard.cell(row, 2).value)
                unit = clean_label(dashboard.cell(row, 3).value)
                code = clean_label(code_raw)

                if is_section_row(row, code, label, unit):
                    title = clean_label(label or code)
                    if not title:
                        continue
                    section_order += 1
                    current_section, _ = FinanceSection.objects.update_or_create(
                        report=report,
                        title=title,
                        defaults={'display_order': section_order},
                    )
                    continue

                if not label:
                    continue

                kind = metric_kind(unit, label)
                scope_type, scope_name = scope_for(current_section.title if current_section else '', label)
                metric_code = f'row_{row}_{code or metric_count + 1}'
                metric, _ = FinanceMetric.objects.update_or_create(
                    report=report,
                    code=metric_code,
                    defaults={
                        'section': current_section,
                        'label': label,
                        'unit': unit,
                        'metric_kind': kind,
                        'scope_type': scope_type,
                        'scope_name': scope_name,
                        'cumulative_behavior': cumulative_behavior(kind, label),
                        'display_order': row,
                        'workbook_sheet': sheet_name,
                        'workbook_row': row,
                        'is_total': label.lower().startswith('total') or 'amount billed water & sewer' in label.lower(),
                        'is_summary': row <= 29,
                    },
                )
                metric_count += 1

                selected_sources = {
                    key: pick_source_row(source_rows[key], source_row_usage[key], label)
                    for key in source_rows
                }
                selected_rows = {key: value[0] for key, value in selected_sources.items()}

                for month_label, (calendar_month, fiscal_index) in MONTHS.items():
                    year = fy_start if calendar_month >= 7 else fy_start + 1
                    target_numeric, target_text = as_decimal(
                        sheet_value(budget, selected_rows['budget'], budget_columns, month_label) if is_current else None
                    )
                    actual_numeric, actual_text = as_decimal(
                        sheet_value(actual, selected_rows['actual'], actual_columns, month_label)
                    )
                    previous_numeric, previous_text = as_decimal(
                        sheet_value(previous, selected_rows['previous'], previous_columns, month_label)
                    )

                    FinanceMonthlyValue.objects.update_or_create(
                        metric=metric,
                        year=year,
                        month=calendar_month,
                        defaults={
                            'fiscal_month_index': fiscal_index,
                            'target_value_numeric': target_numeric,
                            'target_value_text': target_text,
                            'actual_value_numeric': actual_numeric,
                            'actual_value_text': actual_text,
                            'previous_year_actual_numeric': previous_numeric,
                            'previous_year_actual_text': previous_text,
                            'source_row': selected_rows['actual'] or row,
                        },
                    )
                    value_count += 1

                for key in source_row_usage:
                    usage_label = selected_sources[key][1]
                    source_row_usage[key][usage_label] = source_row_usage[key].get(usage_label, 0) + 1

            imported_reports += 1
            imported_metrics += metric_count
            imported_values += value_count
            self.stdout.write(
                self.style.SUCCESS(
                    f'Imported {metric_count} metrics and {value_count} monthly values for {report} from {sheet_name}.'
                )
            )

        self.stdout.write(
            self.style.SUCCESS(
                f'Imported {imported_reports} finance reports, {imported_metrics} metrics and {imported_values} monthly values.'
            )
        )
