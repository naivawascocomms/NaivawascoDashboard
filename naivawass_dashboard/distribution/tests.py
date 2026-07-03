from datetime import date
from decimal import Decimal
import os
import tempfile
from unittest.mock import patch

from django.core.management import call_command
from django.core.exceptions import ValidationError
from django.test import TestCase
from django.contrib.auth import get_user_model
from rest_framework.test import APIClient
from openpyxl import Workbook

from distribution.models import (
    BillingCycle,
    CommercialDashboardKPI,
    CommercialDashboardMonthlyValue,
    CommercialDashboardReport,
    CommercialDashboardSnapshot,
    CustomerBillingData,
    DailyDistribution,
    DistributionMeter,
    DistributionRegion,
    GlobalNRWPerformance,
    MonthlyDistribution,
    RegionalDistribution,
    Zone,
    ZoneBillingCycle,
    ZoneSupplyConfiguration,
    DMA,
)
from distribution.utils import sync_regional_billing_cycles_from_zone_cycles
from metering.models import DistributionWaterMeterAssignment, WaterMeter, WaterMeterReading
from production.models import MonthlyProduction, ProductionSite, Region


class DistributionStructureSyncTests(TestCase):
    def setUp(self):
        self.central = Region.objects.create(code='CENTRAL', name='Central')
        self.south = Region.objects.create(code='SOUTH', name='Southern')
        self.east = Region.objects.create(code='EAST', name='Eastern')

    def test_sync_distribution_structure_creates_workbook_hierarchy(self):
        legacy_region = DistributionRegion.objects.create(code='LEGACY', name='Legacy')
        legacy_zone = Zone.objects.create(region=legacy_region, code='LEGACY-ZONE', name='Legacy Zone')
        meter = DistributionMeter.objects.create(
            meter_location_type='BULK_SUPPLY',
            zone=legacy_zone,
            meter_number='LEGACY-METER-1',
            installation_date=date(2026, 4, 15),
        )

        call_command('sync_distribution_structure')

        self.assertEqual(DistributionRegion.objects.count(), 3)
        self.assertEqual(Zone.objects.count(), 16)

        central_region = DistributionRegion.objects.get(code='CENTRAL')
        self.assertEqual(central_region.dashboard_supply_kpi_code, 'RA')
        self.assertEqual(central_region.production_region.code, 'CENTRAL')

        kayole = Zone.objects.get(code='KAYOLE')
        self.assertEqual(kayole.region.code, 'EAST')
        self.assertEqual(kayole.dashboard_supply_kpi_code, '103 (4)')
        self.assertEqual(kayole.dashboard_billed_kpi_code, '110 (4)')
        self.assertEqual(kayole.dashboard_nrw_pct_kpi_code, '124 (4)')
        self.assertEqual(kayole.supply_aggregation_method, 'UNSET')

        meter.refresh_from_db()
        self.assertIsNone(meter.zone)
        self.assertIsNone(meter.dma)


class DistributionPhaseOneMeterMappingTests(TestCase):
    def setUp(self):
        Region.objects.create(code='CENTRAL', name='Central')
        Region.objects.create(code='SOUTH', name='Southern')
        Region.objects.create(code='EAST', name='Eastern')
        call_command('sync_distribution_structure')

    def test_phase_one_meter_mapping_only_maps_safe_inventory(self):
        lower_site = DistributionMeter.objects.create(
            meter_location_type='BULK_SUPPLY',
            meter_number='LOWER SITE UL1',
            notes='LOWER SITE',
            installation_date=date(2026, 4, 15),
        )
        kayole_1 = DistributionMeter.objects.create(
            meter_location_type='DMA_INLET',
            meter_number='KAYOLE UL1',
            notes='Legacy zone/DMA mapping cleared during workbook structure sync.',
            installation_date=date(2026, 4, 15),
        )
        kayole_2 = DistributionMeter.objects.create(
            meter_location_type='DMA_INLET',
            meter_number='KAYOLE UL2',
            notes='Legacy zone/DMA mapping cleared during workbook structure sync.',
            installation_date=date(2026, 4, 15),
        )
        ambiguous = DistributionMeter.objects.create(
            meter_location_type='BULK_SUPPLY',
            meter_number='WATER WORKS UL1',
            notes='WATER WORKS',
            installation_date=date(2026, 4, 15),
        )

        call_command('map_distribution_meters_phase1')

        lower_site.refresh_from_db()
        self.assertEqual(lower_site.zone.code, 'SITE')
        self.assertIsNone(lower_site.dma)

        kayole_1.refresh_from_db()
        kayole_2.refresh_from_db()
        self.assertEqual(kayole_1.zone.code, 'KAYOLE')
        self.assertEqual(kayole_1.dma.code, 'KAYOLE-DMA')
        self.assertEqual(kayole_2.dma.code, 'KAYOLE-DMA')

        ambiguous.refresh_from_db()
        self.assertIsNone(ambiguous.zone)
        self.assertIsNone(ambiguous.dma)

        self.assertEqual(Zone.objects.get(code='SITE').supply_aggregation_method, 'ZONE_METER')
        self.assertEqual(Zone.objects.get(code='KAYOLE').supply_aggregation_method, 'DMA_SUM')


class ZoneSupplyConfigurationTests(TestCase):
    def setUp(self):
        region = DistributionRegion.objects.create(code='CENTRAL', name='Central')
        self.zone = Zone.objects.create(region=region, code='CBD', name='CBD')
        self.meter = DistributionMeter.objects.create(
            meter_location_type='ZONE_INLET',
            zone=self.zone,
            meter_number='CBD-METER-1',
            installation_date=date(2026, 4, 15),
        )
        self.other_zone = Zone.objects.create(region=region, code='SITE', name='Site')
        self.other_meter = DistributionMeter.objects.create(
            meter_location_type='ZONE_INLET',
            zone=self.other_zone,
            meter_number='SITE-METER-1',
            installation_date=date(2026, 4, 15),
        )

    def test_one_meter_configuration_validates_primary_meter_zone(self):
        config = ZoneSupplyConfiguration(zone=self.zone, aggregation_method='ONE_METER', primary_meter=self.meter)
        config.full_clean()

    def test_one_meter_configuration_allows_shared_meter_from_another_zone(self):
        config = ZoneSupplyConfiguration(zone=self.zone, aggregation_method='ONE_METER', primary_meter=self.other_meter)
        config.full_clean()

    def test_sum_of_dma_configuration_requires_dma_selection(self):
        config = ZoneSupplyConfiguration.objects.create(zone=self.zone, aggregation_method='SUM_OF_DMA_METERS')
        with self.assertRaises(ValidationError):
            config.full_clean()

        dma = DMA.objects.create(zone=self.zone, code='CBD-DMA', name='CBD DMA')
        config.component_dmas.add(dma)
        config.full_clean()

    def test_custom_assignment_configuration_allows_zone_level_assignment_sets(self):
        config = ZoneSupplyConfiguration.objects.create(zone=self.zone, aggregation_method='CUSTOM_ASSIGNMENTS')
        config.full_clean()


class SharedDistributionAssignmentAggregationTests(TestCase):
    def setUp(self):
        self.region = DistributionRegion.objects.create(code='CENTRAL', name='Central')
        self.zone = Zone.objects.create(region=self.region, code='CBD', name='CBD')
        self.zone.supply_configuration = ZoneSupplyConfiguration.objects.create(
            zone=self.zone,
            aggregation_method='CUSTOM_ASSIGNMENTS',
        )
        self.water_meter_a = WaterMeter.objects.create(
            meter_number='A-METER',
            installation_date=date(2026, 1, 1),
        )
        self.water_meter_b = WaterMeter.objects.create(
            meter_number='B-METER',
            installation_date=date(2026, 1, 1),
        )
        self.water_meter_c = WaterMeter.objects.create(
            meter_number='C-METER',
            installation_date=date(2026, 1, 1),
        )
        DistributionWaterMeterAssignment.objects.create(
            water_meter=self.water_meter_a,
            zone=self.zone,
            assignment_role='BULK_SUPPLY',
            allocation_percentage=Decimal('100'),
        )
        DistributionWaterMeterAssignment.objects.create(
            water_meter=self.water_meter_b,
            zone=self.zone,
            assignment_role='BULK_SUPPLY',
            allocation_percentage=Decimal('100'),
        )
        DistributionWaterMeterAssignment.objects.create(
            water_meter=self.water_meter_c,
            zone=self.zone,
            assignment_role='BULK_SUPPLY',
            allocation_percentage=Decimal('-100'),
        )

    def test_custom_assignments_support_signed_meter_contributions(self):
        reading_date = date(2026, 4, 1)
        WaterMeterReading.objects.create(
            water_meter=self.water_meter_a,
            reading_date=reading_date,
            current_reading=Decimal('100'),
            previous_reading=Decimal('0'),
            consumption=Decimal('100'),
            is_validated=True,
        )
        WaterMeterReading.objects.create(
            water_meter=self.water_meter_b,
            reading_date=reading_date,
            current_reading=Decimal('60'),
            previous_reading=Decimal('0'),
            consumption=Decimal('60'),
            is_validated=True,
        )
        WaterMeterReading.objects.create(
            water_meter=self.water_meter_c,
            reading_date=reading_date,
            current_reading=Decimal('25'),
            previous_reading=Decimal('0'),
            consumption=Decimal('25'),
            is_validated=True,
        )

        from distribution.utils import aggregate_daily_distribution

        daily = aggregate_daily_distribution(self.zone, reading_date)

        self.assertIsNotNone(daily)
        self.assertEqual(daily.volume_supplied_m3, Decimal('135'))


class SharedDistributionSignalTests(TestCase):
    def setUp(self):
        self.region = DistributionRegion.objects.create(code='CENTRAL', name='Central')
        self.zone = Zone.objects.create(region=self.region, code='CBD', name='CBD')
        ZoneSupplyConfiguration.objects.create(
            zone=self.zone,
            aggregation_method='CUSTOM_ASSIGNMENTS',
        )
        self.water_meter = WaterMeter.objects.create(
            meter_number='CBD-METER-1',
            display_name='CBD Inlet',
            installation_date=date(2026, 1, 1),
            initial_reading=Decimal('500'),
        )
        DistributionWaterMeterAssignment.objects.create(
            water_meter=self.water_meter,
            zone=self.zone,
            assignment_role='BULK_SUPPLY',
            allocation_percentage=Decimal('100'),
        )

    def test_validated_shared_water_meter_reading_automatically_creates_daily_distribution(self):
        reading_date = date(2026, 4, 2)
        WaterMeterReading.objects.create(
            water_meter=self.water_meter,
            reading_date=reading_date,
            current_reading=Decimal('620'),
            previous_reading=Decimal('500'),
            consumption=Decimal('120'),
            is_validated=True,
        )

        daily = DailyDistribution.objects.get(zone=self.zone, distribution_date=reading_date)
        self.assertEqual(daily.volume_supplied_m3, Decimal('120'))
        self.assertTrue(daily.is_validated)


class CommercialDashboardImportTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = get_user_model().objects.create_user(
            username='commercial-dashboard-tester',
            password='secret123',
        )
        self.client.force_authenticate(self.user)

        self.central = DistributionRegion.objects.create(code='CENTRAL', name='Central')
        self.zone = Zone.objects.create(region=self.central, code='CBD', name='CBD')

    def tearDown(self):
        workbook_path = getattr(self, 'workbook_path', None)
        if workbook_path and os.path.exists(workbook_path):
            os.unlink(workbook_path)

    def _create_workbook(self):
        workbook = Workbook()

        dashboard = workbook.active
        dashboard.title = 'Dashboard'
        dashboard['A1'] = 'SALES & CC DASHBOARD'
        dashboard['B2'] = date(2026, 2, 1)
        dashboard['A5'] = 'NEW CONNECTIONS WATER'
        dashboard['Q5'] = 'monthly'
        dashboard['R5'] = 'cumluative'
        dashboard['A6'] = 1
        dashboard['B6'] = 'CBD - NEW WATER'
        dashboard['C6'] = 'No.'
        dashboard['G6'] = 1
        dashboard['H6'] = 0
        dashboard['I6'] = 0
        dashboard['M6'] = 8
        dashboard['N6'] = 2
        dashboard['O6'] = Decimal('25') / Decimal('100')
        dashboard['A7'] = None
        dashboard['B7'] = 'TOTAL NEW WATER CONNECTIONS CENTRAL REGION'
        dashboard['C7'] = 'No.'
        dashboard['G7'] = 21
        dashboard['H7'] = 7
        dashboard['I7'] = Decimal('33.3333')
        dashboard['M7'] = 167
        dashboard['N7'] = 109
        dashboard['O7'] = Decimal('65.2695')
        dashboard['A8'] = 'BILLING CYCLE'
        dashboard['Q8'] = 'monthly'
        dashboard['R8'] = 'cumluative'
        dashboard['A9'] = 1
        dashboard['B9'] = 'Central'
        dashboard['C9'] = 'Date'
        dashboard['G9'] = '18th-28th'
        dashboard['H9'] = '18th-27th'
        dashboard['M9'] = '18th-28th'
        dashboard['N9'] = '18th-27th'
        dashboard['A10'] = 'Prepared by:'

        budget = workbook.create_sheet('budget 2025-26')
        budget['B4'] = 'MONTH'
        budget['C4'] = 'UNIT'
        month_columns_budget = {
            'JUL': 'D', 'AUG': 'E', 'SEP': 'F', 'OCT': 'G', 'NOV': 'H', 'DEC': 'I',
            'JAN': 'J', 'FEB': 'K', 'MAR': 'L', 'APR': 'M', 'MAY': 'N', 'JUN': 'O',
        }
        for label, column in month_columns_budget.items():
            budget[f'{column}4'] = label
        budget['A5'] = 'NEW CONNECTIONS WATER'
        budget['A6'] = 1
        budget['B6'] = 'CBD - NEW WATER'
        budget['C6'] = 'No.'
        for column in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            budget[f'{column}6'] = 1
        budget['A7'] = 0
        budget['B7'] = 'TOTAL NEW WATER CONNECTIONS CENTRAL REGION'
        budget['C7'] = 'No.'
        totals = [20, 21, 21, 21, 21, 21, 21, 21]
        for column, value in zip(['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K'], totals):
            budget[f'{column}7'] = value
        budget['A8'] = 'BILLING CYCLE'
        budget['A9'] = 1
        budget['B9'] = 'Central'
        budget['C9'] = 'Date'
        budget['K9'] = '18th-28th'

        actual = workbook.create_sheet('Input Actual2025-26')
        actual['B4'] = 'MONTH'
        actual['C4'] = 'UNIT'
        actual['E4'] = 'JUL'
        actual['F4'] = 'AUG'
        actual['G4'] = 'SEP'
        actual['H4'] = 'OCT'
        actual['I4'] = 'NOV'
        actual['J4'] = 'DEC'
        actual['K4'] = 'JAN'
        actual['L4'] = 'FEB'
        actual['M4'] = 'MAR'
        actual['N4'] = 'APR'
        actual['O4'] = 'MAY'
        actual['P4'] = 'JUN'
        actual['A5'] = 'NEW CONNECTIONS WATER'
        actual['A6'] = 1
        actual['B6'] = 'CBD - NEW WATER'
        actual['C6'] = 'No.'
        actual['K6'] = 1
        actual['L6'] = 1
        actual['A7'] = 0
        actual['B7'] = 'TOTAL NEW WATER CONNECTIONS CENTRAL REGION'
        actual['C7'] = 'No.'
        for column, value in zip(['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L'], [5, 13, 7, 23, 6, 14, 34, 7]):
            actual[f'{column}7'] = value
        actual['A8'] = 'BILLING CYCLE'
        actual['A9'] = 1
        actual['B9'] = 'Central'
        actual['C9'] = 'Date'
        actual['L9'] = '18th-27th'

        general = workbook.create_sheet('General')
        general['A3'] = 'Actual month'
        general['B3'] = 8
        general['A4'] = 'Percentage volume sewerage of water'
        general['B4'] = Decimal('0.75')

        with tempfile.NamedTemporaryFile(suffix=' 2025-26.xlsx', delete=False) as handle:
            workbook.save(handle.name)
            self.workbook_path = handle.name

    def test_import_command_creates_report_kpis_values_and_dashboard_payload(self):
        self._create_workbook()

        call_command('import_sales_cc_dashboard', self.workbook_path)

        report = CommercialDashboardReport.objects.get(fiscal_year_label='2025-26')
        self.assertEqual(report.current_snapshot_date, date(2026, 2, 1))
        self.assertEqual(report.current_fiscal_month_index, 8)
        self.assertEqual(report.sewerage_percentage_of_water, Decimal('0.7500'))

        cbd_kpi = CommercialDashboardKPI.objects.get(label='CBD - NEW WATER')
        self.assertEqual(cbd_kpi.scope_type, 'ZONE')
        self.assertEqual(cbd_kpi.zone, self.zone)

        monthly_value = CommercialDashboardMonthlyValue.objects.get(kpi=cbd_kpi, month=2)
        self.assertEqual(monthly_value.target_value_numeric, Decimal('1.0000'))
        self.assertEqual(monthly_value.actual_value_numeric, Decimal('1.0000'))

        snapshot = CommercialDashboardSnapshot.objects.get(kpi=cbd_kpi, snapshot_year=2026, snapshot_month=2)
        self.assertEqual(snapshot.cumulative_target_numeric, Decimal('8.0000'))
        self.assertEqual(snapshot.cumulative_actual_numeric, Decimal('2.0000'))

        response = self.client.get(f'/api/distribution/commercial-dashboard-reports/{report.id}/dashboard/', {'month': 2})
        self.assertEqual(response.status_code, 200)

        sections = response.data['sections']
        self.assertEqual(len(sections), 2)
        first_section_rows = sections[0]['rows']
        cbd_row = next(row for row in first_section_rows if row['label'] == 'CBD - NEW WATER')
        total_row = next(row for row in first_section_rows if row['label'] == 'TOTAL NEW WATER CONNECTIONS CENTRAL REGION')
        date_row = sections[1]['rows'][0]

        self.assertEqual(cbd_row['monthly_target']['numeric'], Decimal('1.0000'))
        self.assertEqual(cbd_row['monthly_actual']['numeric'], Decimal('0.0000'))
        self.assertEqual(cbd_row['cumulative_actual']['numeric'], Decimal('2.0000'))
        self.assertEqual(total_row['scope_type'], 'REGION')
        self.assertEqual(date_row['monthly_target']['raw'], '18th-28th')
        self.assertEqual(date_row['monthly_actual']['raw'], '18th-27th')


class BillingCycleWorkflowTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = get_user_model().objects.create_user(
            username='billing-tester',
            password='secret123',
        )
        self.client.force_authenticate(self.user)

        self.region = DistributionRegion.objects.create(code='CENTRAL', name='Central')
        self.zone_a = Zone.objects.create(region=self.region, code='CBD', name='CBD')
        self.zone_b = Zone.objects.create(region=self.region, code='SITE', name='Site')

    def test_zone_billing_cycle_rejects_overlapping_dates(self):
        ZoneBillingCycle.objects.create(
            zone=self.zone_a,
            year=2026,
            month=3,
            opening_date=date(2026, 3, 1),
            closing_date=date(2026, 3, 31),
        )

        overlapping = ZoneBillingCycle(
            zone=self.zone_a,
            year=2026,
            month=4,
            opening_date=date(2026, 3, 20),
            closing_date=date(2026, 4, 20),
        )

        with self.assertRaises(ValidationError):
            overlapping.full_clean()

    def test_sync_regional_billing_cycles_uses_min_and_max_zone_dates(self):
        ZoneBillingCycle.objects.create(
            zone=self.zone_a,
            year=2026,
            month=3,
            opening_date=date(2026, 2, 26),
            closing_date=date(2026, 3, 27),
            is_finalized=False,
        )
        ZoneBillingCycle.objects.create(
            zone=self.zone_b,
            year=2026,
            month=3,
            opening_date=date(2026, 3, 1),
            closing_date=date(2026, 3, 30),
            is_finalized=True,
        )

        sync_regional_billing_cycles_from_zone_cycles(year=2026, month=3, region=self.region)

        cycle = BillingCycle.objects.get(region=self.region, year=2026, month=3)
        self.assertEqual(cycle.start_date, date(2026, 2, 26))
        self.assertEqual(cycle.end_date, date(2026, 3, 30))
        self.assertTrue(cycle.is_finalized)

    def test_recalculate_month_builds_monthly_and_regional_distribution(self):
        zone_cycle_a = ZoneBillingCycle.objects.create(
            zone=self.zone_a,
            year=2026,
            month=3,
            opening_date=date(2026, 3, 1),
            closing_date=date(2026, 3, 31),
        )
        zone_cycle_b = ZoneBillingCycle.objects.create(
            zone=self.zone_b,
            year=2026,
            month=3,
            opening_date=date(2026, 3, 2),
            closing_date=date(2026, 3, 29),
        )

        sync_regional_billing_cycles_from_zone_cycles(year=2026, month=3, region=self.region)
        cycle = BillingCycle.objects.get(region=self.region, year=2026, month=3)

        DailyDistribution.objects.create(
            zone=self.zone_a,
            distribution_date=date(2026, 3, 5),
            volume_supplied_m3=Decimal('100'),
            is_validated=True,
        )
        DailyDistribution.objects.create(
            zone=self.zone_a,
            distribution_date=date(2026, 3, 20),
            volume_supplied_m3=Decimal('50'),
            is_validated=True,
        )
        DailyDistribution.objects.create(
            zone=self.zone_b,
            distribution_date=date(2026, 3, 10),
            volume_supplied_m3=Decimal('80'),
            is_validated=True,
        )

        CustomerBillingData.objects.create(
            zone=self.zone_a,
            zone_billing_cycle=zone_cycle_a,
            total_volume_billed_m3=Decimal('120'),
            number_of_bills_generated=10,
            number_of_active_connections=10,
            water_revenue=Decimal('1200'),
            sewer_revenue=Decimal('300'),
        )
        CustomerBillingData.objects.create(
            zone=self.zone_b,
            zone_billing_cycle=zone_cycle_b,
            total_volume_billed_m3=Decimal('60'),
            number_of_bills_generated=8,
            number_of_active_connections=8,
            water_revenue=Decimal('600'),
            sewer_revenue=Decimal('150'),
        )

        response = self.client.post(
            '/api/distribution/zone-billing-cycles/recalculate_month/',
            {'year': 2026, 'month': 3, 'region': self.region.id},
            format='json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['zones_recalculated'], 2)
        self.assertEqual(response.data['regions_recalculated'], 1)

        monthly_a = MonthlyDistribution.objects.get(zone=self.zone_a, billing_cycle=cycle)
        monthly_b = MonthlyDistribution.objects.get(zone=self.zone_b, billing_cycle=cycle)
        regional = RegionalDistribution.objects.get(region=self.region, billing_cycle=cycle)

        self.assertEqual(monthly_a.zone_billing_cycle_id, zone_cycle_a.id)
        self.assertEqual(monthly_b.zone_billing_cycle_id, zone_cycle_b.id)
        self.assertEqual(monthly_a.volume_supplied_m3, Decimal('150'))
        self.assertEqual(monthly_a.volume_billed_m3, Decimal('120'))
        self.assertEqual(monthly_a.nrw_m3, Decimal('30'))

        self.assertEqual(monthly_b.volume_supplied_m3, Decimal('80'))
        self.assertEqual(monthly_b.volume_billed_m3, Decimal('60'))
        self.assertEqual(monthly_b.nrw_m3, Decimal('20'))

        self.assertEqual(regional.volume_supplied_m3, Decimal('230'))
        self.assertEqual(regional.volume_billed_m3, Decimal('180'))
        self.assertEqual(regional.nrw_m3, Decimal('50'))
        self.assertEqual(regional.amount_billed_water, Decimal('1800'))
        self.assertEqual(regional.amount_billed_sewer, Decimal('450'))
        self.assertEqual(regional.active_water_connections, 18)

    def test_billing_cycle_create_builds_commercial_summaries(self):
        DailyDistribution.objects.create(
            zone=self.zone_a,
            distribution_date=date(2026, 4, 3),
            volume_supplied_m3=Decimal('100'),
            is_validated=True,
        )
        DailyDistribution.objects.create(
            zone=self.zone_b,
            distribution_date=date(2026, 4, 4),
            volume_supplied_m3=Decimal('80'),
            is_validated=True,
        )

        response = self.client.post(
            '/api/distribution/billing-cycles/',
            {
                'region': self.region.id,
                'year': 2026,
                'month': 4,
                'start_date': '2026-04-01',
                'end_date': '2026-04-30',
                'is_finalized': True,
            },
            format='json',
        )

        self.assertEqual(response.status_code, 201)
        cycle = BillingCycle.objects.get(region=self.region, year=2026, month=4)
        monthly_a = MonthlyDistribution.objects.get(zone=self.zone_a, billing_cycle=cycle)
        monthly_b = MonthlyDistribution.objects.get(zone=self.zone_b, billing_cycle=cycle)
        regional = RegionalDistribution.objects.get(region=self.region, billing_cycle=cycle)
        global_nrw = GlobalNRWPerformance.objects.get(billing_cycle=cycle)

        self.assertEqual(monthly_a.volume_supplied_m3, Decimal('100'))
        self.assertEqual(monthly_b.volume_supplied_m3, Decimal('80'))
        self.assertEqual(monthly_a.volume_billed_m3, Decimal('0'))
        self.assertEqual(regional.volume_supplied_m3, Decimal('180'))
        self.assertEqual(regional.volume_billed_m3, Decimal('0'))
        self.assertEqual(global_nrw.volume_billed_to_customers_m3, Decimal('0'))

    @patch('distribution.views.timezone.localdate', return_value=date(2026, 3, 15))
    def test_current_cycle_returns_cycle_covering_today(self, _mock_localdate):
        BillingCycle.objects.create(
            region=self.region,
            year=2026,
            month=4,
            start_date=date(2026, 4, 1),
            end_date=date(2026, 4, 30),
        )
        active_cycle = BillingCycle.objects.create(
            region=self.region,
            year=2026,
            month=3,
            start_date=date(2026, 3, 1),
            end_date=date(2026, 3, 31),
        )

        response = self.client.get(
            f'/api/distribution/billing-cycles/current_cycle/?region={self.region.id}'
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['id'], active_cycle.id)
        self.assertEqual(response.data['month'], 3)
