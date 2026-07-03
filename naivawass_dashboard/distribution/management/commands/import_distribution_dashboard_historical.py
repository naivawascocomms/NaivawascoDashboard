from calendar import monthrange
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand
from django.db import transaction

from distribution.models import (
    BillingCycle,
    CustomerBillingData,
    DistributionRegion,
    GlobalNRWPerformance,
    MonthlyDistribution,
    RegionalDistribution,
    Zone,
)

import openpyxl


ACTUAL_SHEETS = [
    'Actual2021',
    'Actual2122',
    'Actual2223',
    'Actuals2425',
    'Actuals2526',
]


ZONE_LABELS = {
    'ZONE - CBD': 'CBD',
    'ZONE - CCCR': 'CCCR',
    'ZONE - LAKEVIEW': 'LAKEVIEW',
    'ZONE - KABATI': 'KABATI',
    'ZONE - SITE AND SERVICES': 'SITE',
    'ZONE - HOPEWELL': 'HOPEWELL',
    'ZONE - KAMERE': 'KAMERE',
    'ZONE - HELLS GATE': 'HELLSGATE',
    'KIHOTO': 'KIHOTO',
    'KAYOLE': 'KAYOLE',
    'MAI-MAHIU': 'MAIMAHIU',
    'KINUNGI': 'KINUNGI',
    'GONDI': 'GONDI',
    'NYONJORO': 'NYONJORO',
    'LONGONOT': 'LONGONOT',
    'IHINDU': 'IHINDU',
}


REGION_HEADER_MAP = {
    'CENTRAL REGION': 'CENTRAL',
    'DTI REGION': 'CENTRAL',
    'SOUTHERN REGION': 'SOUTH',
    'EASTERN REGION': 'EAST',
}


def normalize_text(value):
    if value is None:
        return ''
    return ' '.join(str(value).strip().upper().replace('—', '-').replace('–', '-').split())


def coerce_decimal(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text or text.startswith('#'):
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def parse_month_headers(ws):
    months = []
    for col in range(4, 16):
        raw = ws.cell(4, col).value
        if isinstance(raw, datetime):
            months.append(date(raw.year, raw.month, 1))
            continue
        text = normalize_text(raw).lower()
        month_lookup = {
            'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
            'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
        }
        if text not in month_lookup:
            months.append(None)
            continue
        year_hint = 2020
        if ws.title == 'Actual2021':
            year_hint = 2020 if month_lookup[text] >= 7 else 2021
        elif ws.title == 'Actual2122':
            year_hint = 2021 if month_lookup[text] >= 7 else 2022
        months.append(date(year_hint, month_lookup[text], 1))
    return months


class Command(BaseCommand):
    help = 'Import historical distribution dashboard actuals up to a cutoff month, using workbook totals rather than meter readings.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            default=str(Path.cwd().parent / 'DISTRIBUTION DASHBOARD.xlsx'),
            help='Path to DISTRIBUTION DASHBOARD.xlsx',
        )
        parser.add_argument('--cutoff-year', type=int, default=2026)
        parser.add_argument('--cutoff-month', type=int, default=2)
        parser.add_argument(
            '--reset',
            action='store_true',
            help='Delete imported historical distribution summaries up to the cutoff before reimporting.',
        )

    @transaction.atomic
    def handle(self, *args, **options):
        workbook_path = Path(options['file'])
        cutoff = date(options['cutoff_year'], options['cutoff_month'], 1)

        if options['reset']:
            self._reset_existing(cutoff)

        wb = openpyxl.load_workbook(workbook_path, data_only=True)

        zone_map = {zone.code: zone for zone in Zone.objects.select_related('region')}
        region_map = {region.code: region for region in DistributionRegion.objects.all()}

        imported_zone_records = 0
        imported_global_records = 0

        for sheet_name in ACTUAL_SHEETS:
            ws = wb[sheet_name]
            months = parse_month_headers(ws)
            parsed = self._parse_sheet(ws, months, cutoff)

            for month_start, zone_values in parsed['zones'].items():
                for zone_code, metrics in zone_values.items():
                    zone = zone_map.get(zone_code)
                    if not zone:
                        continue
                    cycle = self._get_or_create_billing_cycle(zone.region, month_start)
                    billed = metrics.get('billed') or Decimal('0')
                    supplied = metrics.get('supplied') or Decimal('0')
                    water_revenue = metrics.get('water_revenue') or Decimal('0')
                    sewer_revenue = metrics.get('sewer_revenue') or Decimal('0')
                    CustomerBillingData.objects.update_or_create(
                        zone=zone,
                        billing_cycle=cycle,
                        defaults={
                            'total_volume_billed_m3': billed,
                            'water_revenue': water_revenue,
                            'sewer_revenue': sewer_revenue,
                            'total_revenue': water_revenue + sewer_revenue,
                            'notes': 'Imported from historical distribution dashboard workbook.',
                        }
                    )
                    MonthlyDistribution.objects.update_or_create(
                        zone=zone,
                        billing_cycle=cycle,
                        defaults={
                            'volume_supplied_m3': supplied,
                            'volume_billed_m3': billed,
                            'notes': 'Imported from historical distribution dashboard workbook.',
                            'is_finalized': True,
                        }
                    )
                    imported_zone_records += 1

            for month_start, metrics in parsed['global'].items():
                cycle = self._get_or_create_global_cycle(month_start, region_map)
                GlobalNRWPerformance.objects.update_or_create(
                    billing_cycle=cycle,
                    defaults={
                        'water_available_for_sale_m3': metrics.get('available') or Decimal('0'),
                        'volume_billed_to_customers_m3': metrics.get('billed') or Decimal('0'),
                        'active_water_connections': int(metrics.get('active_water_connections') or 0),
                        'active_sewer_connections': int(metrics.get('active_sewer_connections') or 0),
                        'inactive_water_connections': int(metrics.get('inactive_water_connections') or 0),
                        'inactive_sewer_connections': int(metrics.get('inactive_sewer_connections') or 0),
                        'total_connections': int(metrics.get('total_connections') or 0),
                        'maintenance_repair_operational_cost': metrics.get('maintenance_cost') or Decimal('0'),
                        'notes': 'Imported from historical distribution dashboard workbook.',
                    }
                )
                imported_global_records += 1

        # Derive regional summaries from imported zone monthly records so region totals remain consistent with zones.
        for cycle in BillingCycle.objects.filter(
            year__lt=cutoff.year
        ) | BillingCycle.objects.filter(year=cutoff.year, month__lte=cutoff.month):
            queryset = MonthlyDistribution.objects.filter(billing_cycle=cycle)
            if not queryset.exists():
                continue
            for region in region_map.values():
                region_zone_records = queryset.filter(zone__region=region)
                if not region_zone_records.exists():
                    continue
                supplied = sum((record.volume_supplied_m3 or Decimal('0')) for record in region_zone_records)
                billed = sum((record.volume_billed_m3 or Decimal('0')) for record in region_zone_records)
                RegionalDistribution.objects.update_or_create(
                    region=region,
                    billing_cycle=cycle,
                    defaults={
                        'volume_supplied_m3': supplied,
                        'volume_billed_m3': billed,
                        'amount_billed_water': sum((record.water_revenue or Decimal('0')) for record in CustomerBillingData.objects.filter(zone__region=region, billing_cycle=cycle)),
                        'amount_billed_sewer': sum((record.sewer_revenue or Decimal('0')) for record in CustomerBillingData.objects.filter(zone__region=region, billing_cycle=cycle)),
                        'active_water_connections': sum(record.number_of_active_connections for record in CustomerBillingData.objects.filter(zone__region=region, billing_cycle=cycle)),
                        'notes': 'Derived from imported historical zone dashboard actuals.',
                        'is_finalized': True,
                    }
                )

        self.stdout.write(self.style.SUCCESS(
            f'Imported historical distribution dashboard actuals up to {cutoff:%Y-%m}. '
            f'Updated {imported_zone_records} zone-month records and {imported_global_records} global records.'
        ))

    def _parse_sheet(self, ws, months, cutoff):
        parsed = {'zones': {}, 'global': {}}
        current_zone_code = None

        for row in range(1, ws.max_row + 1):
            label = normalize_text(ws.cell(row, 2).value)
            unit = normalize_text(ws.cell(row, 3).value)

            if label in ZONE_LABELS:
                current_zone_code = ZONE_LABELS[label]
                continue

            metric_kind = None
            if 'WATER AVAILABLE FOR SALES' in label:
                metric_kind = ('global', 'available')
            elif 'VOLUME WATER BILLED' in label and current_zone_code is None:
                metric_kind = ('global', 'billed')
            elif label == 'GLOBAL NRW':
                metric_kind = ('global', 'nrw_pct')
            elif 'NUMBER OF ACTIVE WATER CONNECTIONS' in label and current_zone_code is None:
                metric_kind = ('global', 'active_water_connections')
            elif 'NUMBER OF ACTIVE SEWER CONNECTIONS' in label and current_zone_code is None:
                metric_kind = ('global', 'active_sewer_connections')
            elif 'NUMBER OF IN-ACTIVE SEWER CONNECTIONS' in label and current_zone_code is None:
                metric_kind = ('global', 'inactive_sewer_connections')
            elif 'NUMBER OF IN-ACTIVE WATER CONNECTIONS' in label and current_zone_code is None:
                metric_kind = ('global', 'inactive_water_connections')
            elif 'NUMBER OF CONNECTIONS' in label and current_zone_code is None:
                metric_kind = ('global', 'total_connections')
            elif 'MAINTENANCE, REPAIR AND OPERATIONAL COST' in label and current_zone_code is None:
                metric_kind = ('global', 'maintenance_cost')
            elif current_zone_code:
                if 'VOLUME WATER SUPPLIED' in label:
                    metric_kind = ('zone', 'supplied')
                elif 'VOLUME WATER SOLD' in label or 'VOLUME WATER BILLED' in label:
                    metric_kind = ('zone', 'billed')
                elif label == 'NRW' and unit == '%':
                    metric_kind = ('zone', 'nrw_pct')
                elif 'AMOUNT BILLED WATER' in label:
                    metric_kind = ('zone', 'water_revenue')
                elif 'AMOUNT BILLED SEWER' in label:
                    metric_kind = ('zone', 'sewer_revenue')

            if metric_kind is None:
                continue

            for idx, month_start in enumerate(months, start=4):
                if not month_start or month_start > cutoff:
                    continue
                value = coerce_decimal(ws.cell(row, idx).value)
                if value is None:
                    continue
                if metric_kind[0] == 'global':
                    parsed['global'].setdefault(month_start, {})[metric_kind[1]] = value
                else:
                    parsed['zones'].setdefault(month_start, {}).setdefault(current_zone_code, {})[metric_kind[1]] = value

        return parsed

    def _get_or_create_billing_cycle(self, region, month_start):
        end_day = monthrange(month_start.year, month_start.month)[1]
        cycle, _ = BillingCycle.objects.get_or_create(
            region=region,
            year=month_start.year,
            month=month_start.month,
            defaults={
                'start_date': month_start,
                'end_date': date(month_start.year, month_start.month, end_day),
                'is_finalized': True,
                'notes': 'Imported historical dashboard billing cycle.',
            }
        )
        return cycle

    def _get_or_create_global_cycle(self, month_start, region_map):
        # Use Central region cycle as the canonical billing cycle anchor for global metrics.
        return self._get_or_create_billing_cycle(region_map['CENTRAL'], month_start)

    def _reset_existing(self, cutoff):
        cycles = BillingCycle.objects.filter(
            year__lt=cutoff.year
        ) | BillingCycle.objects.filter(year=cutoff.year, month__lte=cutoff.month)
        GlobalNRWPerformance.objects.filter(billing_cycle__in=cycles).delete()
        RegionalDistribution.objects.filter(billing_cycle__in=cycles).delete()
        MonthlyDistribution.objects.filter(billing_cycle__in=cycles).delete()
        CustomerBillingData.objects.filter(billing_cycle__in=cycles).delete()
        BillingCycle.objects.filter(id__in=cycles.values_list('id', flat=True)).delete()
