import os
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from distribution.models import (
    CommercialDashboardKPI,
    CommercialDashboardMonthlyValue,
    CommercialDashboardReport,
    CommercialDashboardSection,
    CommercialDashboardSnapshot,
    DistributionRegion,
    Zone,
)


FY_MONTH_ORDER = [7, 8, 9, 10, 11, 12, 1, 2, 3, 4, 5, 6]
MONTH_NAME_TO_NUMBER = {
    'JUL': 7,
    'AUG': 8,
    'SEP': 9,
    'OCT': 10,
    'NOV': 11,
    'DEC': 12,
    'JAN': 1,
    'FEB': 2,
    'MAR': 3,
    'APR': 4,
    'MAY': 5,
    'JUN': 6,
}
ZONE_SCOPE_ALIASES = {
    'CBD': 'CBD',
    'CCCR': 'CCCR',
    'KABATI': 'KABATI',
    'SITE': 'SITE',
    'SITEANDSERVICES': 'SITE',
    'KIHOTO': 'KIHOTO',
    'LAKEVIEW': 'LAKEVIEW',
    'HOPEWELL': 'HOPEWELL',
    'HELLSGATE': 'HELLSGATE',
    'KAMERE': 'KAMERE',
    'MAIMAHIU': 'MAIMAHIU',
    'MAIMAHIUU': 'MAIMAHIU',
    'MAIMAHIUH': 'MAIMAHIU',
    'MAIMAHIUHNEWWATER': 'MAIMAHIU',
    'MAIMAHIU': 'MAIMAHIU',
    'LONGONOT': 'LONGONOT',
    'KAYOLE': 'KAYOLE',
    'IHINDU': 'IHINDU',
}
REGION_SCOPE_ALIASES = {
    'CENTRAL': 'CENTRAL',
    'CENTRALREGION': 'CENTRAL',
    'SOUTHERN': 'SOUTH',
    'SOUTHERNREGION': 'SOUTH',
    'EASTERN': 'EAST',
    'EASTERNREGION': 'EAST',
}


def normalize_text(value):
    return re.sub(r'[^A-Z0-9]+', '', str(value or '').upper())


def parse_numeric(value):
    if value is None or value == '':
        return None
    if isinstance(value, bool):
        return Decimal(int(value))
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    if isinstance(value, (date, datetime)):
        return None
    cleaned = str(value).strip()
    if not cleaned or cleaned.lower() == 'l':
        return None
    cleaned = cleaned.replace(',', '')
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def parse_text(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float, Decimal)):
        return ''
    cleaned = str(value).strip()
    if cleaned.lower() == 'l':
        return ''
    return cleaned


class Command(BaseCommand):
    help = 'Import the Sales & Customer Care dashboard workbook into the distribution commercial dashboard backend.'

    def add_arguments(self, parser):
        parser.add_argument(
            'workbook',
            nargs='?',
            default=None,
            help='Path to SALES  CC DASHBOARD workbook. Defaults to ../SALES  CC DASHBOARD 2025-26.xlsx relative to manage.py.',
        )
        parser.add_argument(
            '--deactivate-others',
            action='store_true',
            help='Mark other commercial dashboard reports as inactive after import.',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Validate and report import results without writing to the database.',
        )

    @transaction.atomic
    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError('openpyxl is required to import the sales dashboard workbook.') from exc

        workbook_path = options['workbook'] or os.path.normpath(
            os.path.join(os.path.dirname(os.path.abspath('manage.py')), '..', 'SALES  CC DASHBOARD 2025-26.xlsx')
        )
        if not os.path.exists(workbook_path):
            raise CommandError(f'Workbook not found: {workbook_path}')

        workbook = openpyxl.load_workbook(workbook_path, data_only=True)
        dashboard_sheet = workbook['Dashboard']
        actual_sheet = workbook['Input Actual2025-26']
        budget_sheet = workbook['budget 2025-26']
        general_sheet = workbook['General']

        fiscal_year_start = self._parse_fiscal_year_start(workbook_path)
        fiscal_year_label = f'{fiscal_year_start}-{str(fiscal_year_start + 1)[-2:]}'
        snapshot_date = dashboard_sheet['B2'].value
        if isinstance(snapshot_date, datetime):
            snapshot_date = snapshot_date.date()

        fiscal_month_index = general_sheet['B3'].value
        fiscal_month_index = int(fiscal_month_index) if fiscal_month_index is not None else None
        sewer_ratio = parse_numeric(general_sheet['B4'].value) or Decimal('0')

        zone_lookup = self._zone_lookup()
        region_lookup = self._region_lookup()
        actual_rows = self._build_month_row_map(actual_sheet)
        budget_rows = self._build_month_row_map(budget_sheet)
        dashboard_rows = self._parse_dashboard_rows(dashboard_sheet)

        report, _ = CommercialDashboardReport.objects.get_or_create(
            fiscal_year_label=fiscal_year_label,
            defaults={
                'name': 'Sales & CC Dashboard',
                'fiscal_year_start': fiscal_year_start,
            },
        )

        report.name = 'Sales & CC Dashboard'
        report.fiscal_year_start = fiscal_year_start
        report.current_snapshot_date = snapshot_date
        report.current_fiscal_month_index = fiscal_month_index
        report.sewerage_percentage_of_water = sewer_ratio
        report.source_workbook = os.path.basename(workbook_path)
        report.is_active = True
        if not options['dry_run']:
            report.save()
            report.kpis.all().delete()
            report.sections.all().delete()
            if options['deactivate_others']:
                CommercialDashboardReport.objects.exclude(pk=report.pk).update(is_active=False)

        section_count = 0
        kpi_count = 0
        monthly_value_count = 0
        snapshot_count = 0

        section_instances = {}
        for row in dashboard_rows:
            section_title = row['section_title']
            section = section_instances.get(section_title)
            if section is None:
                section_count += 1
                section = CommercialDashboardSection(
                    report=report,
                    title=section_title,
                    display_order=section_count,
                    workbook_row=row['section_row'],
                )
                if not options['dry_run']:
                    section.save()
                section_instances[section_title] = section

            scope_type, region, zone = self._resolve_scope(row['label'], zone_lookup, region_lookup)
            actual_row = actual_rows.get(normalize_text(row['label']))
            budget_row = budget_rows.get(normalize_text(row['label']))

            kpi_count += 1
            kpi = CommercialDashboardKPI(
                report=report,
                section=section,
                label=row['label'],
                unit=row['unit'],
                item_number=row['item_number'],
                subgroup_title=row['subgroup_title'],
                scope_type=scope_type,
                region=region,
                zone=zone,
                display_order=kpi_count,
                workbook_row=row['workbook_row'],
                is_total=row['is_total'],
                is_summary=row['is_summary'],
                is_percentage='%' in str(row['unit'] or ''),
            )
            if not options['dry_run']:
                kpi.save()

            for month in FY_MONTH_ORDER:
                target_source = budget_row['values'].get(month) if budget_row else None
                actual_source = actual_row['values'].get(month) if actual_row else None
                target_numeric = parse_numeric(target_source)
                actual_numeric = parse_numeric(actual_source)
                target_text = parse_text(target_source)
                actual_text = parse_text(actual_source)
                if (
                    target_numeric is None and actual_numeric is None and
                    not target_text and not actual_text
                ):
                    continue

                monthly_value_count += 1
                if not options['dry_run']:
                    CommercialDashboardMonthlyValue.objects.create(
                        kpi=kpi,
                        month=month,
                        target_value_numeric=target_numeric,
                        target_value_text=target_text,
                        actual_value_numeric=actual_numeric,
                        actual_value_text=actual_text,
                    )

            if snapshot_date:
                snapshot_count += 1
                if not options['dry_run']:
                    CommercialDashboardSnapshot.objects.create(
                        kpi=kpi,
                        snapshot_year=snapshot_date.year,
                        snapshot_month=snapshot_date.month,
                        monthly_target_numeric=parse_numeric(row['monthly_target']),
                        monthly_target_text=parse_text(row['monthly_target']),
                        monthly_actual_numeric=parse_numeric(row['monthly_actual']),
                        monthly_actual_text=parse_text(row['monthly_actual']),
                        monthly_realization_percent=parse_numeric(row['monthly_realization']),
                        cumulative_target_numeric=parse_numeric(row['cumulative_target']),
                        cumulative_target_text=parse_text(row['cumulative_target']),
                        cumulative_actual_numeric=parse_numeric(row['cumulative_actual']),
                        cumulative_actual_text=parse_text(row['cumulative_actual']),
                        cumulative_realization_percent=parse_numeric(row['cumulative_realization']),
                    )

        if options['dry_run']:
            transaction.set_rollback(True)

        self.stdout.write(self.style.SUCCESS(
            f'Imported commercial dashboard structure for {fiscal_year_label}. '
            f'Sections: {section_count}, KPIs: {kpi_count}, monthly values: {monthly_value_count}, snapshots: {snapshot_count}.'
        ))

    def _parse_fiscal_year_start(self, workbook_path):
        match = re.search(r'(\d{4})-(\d{2})', os.path.basename(workbook_path))
        if match:
            return int(match.group(1))
        raise CommandError('Could not derive fiscal year start from workbook filename.')

    def _build_month_row_map(self, worksheet):
        month_columns = {}
        for column in range(1, worksheet.max_column + 1):
            month_label = normalize_text(worksheet.cell(row=4, column=column).value)
            if month_label in MONTH_NAME_TO_NUMBER:
                month_columns[MONTH_NAME_TO_NUMBER[month_label]] = column

        row_map = {}
        for row in range(5, worksheet.max_row + 1):
            label = worksheet.cell(row=row, column=2).value
            unit = worksheet.cell(row=row, column=3).value
            if not label or not unit:
                continue
            if normalize_text(label) == 'MONTH':
                continue
            row_map[normalize_text(label)] = {
                'label': str(label).strip(),
                'unit': str(unit).strip(),
                'values': {
                    month: worksheet.cell(row=row, column=column).value
                    for month, column in month_columns.items()
                },
            }
        return row_map

    def _parse_dashboard_rows(self, worksheet):
        rows = []
        current_section = None
        current_section_row = None
        current_subgroup = ''

        for row in range(5, worksheet.max_row + 1):
            section_candidate = worksheet.cell(row=row, column=1).value
            label = worksheet.cell(row=row, column=2).value
            unit = worksheet.cell(row=row, column=3).value

            if section_candidate and str(section_candidate).strip().upper().startswith('PREPARED BY'):
                break

            if label and unit:
                rows.append({
                    'section_title': current_section or 'General',
                    'section_row': current_section_row,
                    'subgroup_title': current_subgroup,
                    'item_number': '' if section_candidate is None else str(section_candidate).strip(),
                    'label': str(label).strip(),
                    'unit': str(unit).strip(),
                    'workbook_row': row,
                    'monthly_target': worksheet.cell(row=row, column=7).value,
                    'monthly_actual': worksheet.cell(row=row, column=8).value,
                    'monthly_realization': worksheet.cell(row=row, column=9).value,
                    'cumulative_target': worksheet.cell(row=row, column=13).value,
                    'cumulative_actual': worksheet.cell(row=row, column=14).value,
                    'cumulative_realization': worksheet.cell(row=row, column=15).value,
                    'is_total': normalize_text(label).startswith('TOTAL'),
                    'is_summary': normalize_text(label).startswith('SUMMARY'),
                })
                continue

            if section_candidate and not label:
                title = str(section_candidate).strip()
                normalized = normalize_text(title)
                if normalized == 'SALESCCDASHBOARD':
                    continue
                if normalized.endswith('REGION') and not normalized.startswith('TOTAL'):
                    current_subgroup = title
                else:
                    current_section = title
                    current_section_row = row
                    current_subgroup = ''

        return rows

    def _zone_lookup(self):
        lookup = {}
        for zone in Zone.objects.select_related('region'):
            lookup[zone.code] = zone
            lookup[normalize_text(zone.name)] = zone
        return lookup

    def _region_lookup(self):
        lookup = {}
        for region in DistributionRegion.objects.all():
            lookup[region.code] = region
            lookup[normalize_text(region.name)] = region
        return lookup

    def _resolve_scope(self, label, zone_lookup, region_lookup):
        normalized_label = normalize_text(label)

        prefix = str(label).split('-', 1)[0].strip()
        normalized_prefix = normalize_text(prefix)
        zone_key = ZONE_SCOPE_ALIASES.get(normalized_prefix)
        zone = zone_lookup.get(zone_key) if zone_key else None
        if zone is not None:
            return 'ZONE', zone.region, zone

        for token, region_code in REGION_SCOPE_ALIASES.items():
            if token in normalized_label:
                region = region_lookup.get(region_code)
                if region is not None:
                    return 'REGION', region, None

        return 'GLOBAL', None, None
