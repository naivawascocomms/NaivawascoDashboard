import os
import re

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from distribution.models import Zone, ZoneSupplyConfiguration
from metering.models import DistributionWaterMeterAssignment, WaterMeter
from metering.sync import refresh_water_meter_display_name
from metering.workbook_names import slugify_meter_name, split_meter_name_and_notes


def normalize_text(value):
    return re.sub(r'[^A-Z0-9]+', '', str(value or '').upper())


WORKBOOK_ZONE_ALIASES = {
    'CBDCBDHASAUNIQUEWAYTOCALCULATEITISTHETOTALOFALLSUPPLYMETERSCCCRMETER': 'CBD',
    'CCCR': 'CCCR',
    'LAKEVIEW': 'LAKEVIEW',
    'KAYOLE': 'KAYOLE',
    'IHINDU': 'IHINDU',
    'NGONDI': 'GONDI',
    'HOPEWELL': 'HOPEWELL',
    'NYONJORO': 'NYONJORO',
    'KAMERE': 'KAMERE',
    'KABATI': 'KABATI',
    'LONGONOT': 'LONGONOT',
}


ZONE_SYNC_PLAN = {
    'HELLSGATE': {
        'distribution_column': 'HELLSGATE',
        'description': 'Mapped from Distribution Zones workbook to the Hells Gate/Kijabe offtake inlet meter.',
        'calculation_notes': 'Workbook sync using the Kijabe offtake meter listed under Hells Gate.',
        'assignments': [
            {
                'distribution_label': 'Kijable Offtake Meter (meter over registering)',
                'meter_display_name': 'KIJABE OFFTAKE METER',
                'allocation_percentage': 100,
            },
        ],
    },
    'CBD': {
        'distribution_column': 'CBD',
        'description': (
            'Mapped from Distribution Zones workbook as a custom shared-meter calculation: '
            'Police Line BH meter + DTI inlet meter + AIC supply meter + Consolata meter - CCCR meter.'
        ),
        'calculation_notes': (
            'Shared production/distribution workbook sync. CBD uses a signed custom assignment set so CCCR is subtracted '
            'from the gross central supply stack, matching the workbook note.'
        ),
        'assignments': [
            {
                'distribution_label': 'Police Line BH meter',
                'meter_display_name': 'Police Line BH meter',
                'production_sheet': 'POLICE LINE',
                'production_label': 'BH1 METER',
                'allocation_percentage': 100,
            },
            {
                'distribution_label': 'DTI Inlet Meter',
                'meter_display_name': 'DTI Inlet Meter',
                'production_sheet': 'DTI',
                'production_label': 'DTI supply meter',
                'allocation_percentage': 100,
            },
            {
                'distribution_label': 'AIC Supply Meter',
                'meter_display_name': 'AIC Supply Meter',
                'production_sheet': 'AIC',
                'production_label': 'BH1 METER',
                'allocation_percentage': 100,
            },
            {
                'distribution_label': 'Consolata Meter',
                'meter_display_name': 'Consolata Meter',
                'allocation_percentage': 100,
            },
            {
                'distribution_label': 'CCCR METER',
                'meter_display_name': 'CCCR MAIN SUPPLY METER',
                'distribution_column': 'CCCR',
                'allocation_percentage': -100,
            },
        ],
    },
    'CCCR': {
        'distribution_column': 'CCCR',
        'description': 'Mapped from Distribution Zones workbook as a dedicated one-meter commercial inlet.',
        'calculation_notes': 'Shared distribution workbook sync using the canonical CCCR meter.',
        'assignments': [
            {
                'distribution_label': 'CCCR METER',
                'meter_display_name': 'CCCR MAIN SUPPLY METER',
                'allocation_percentage': 100,
            },
        ],
    },
    'LAKEVIEW': {
        'distribution_column': 'LAKEVIEW',
        'description': (
            'Mapped from Distribution Zones workbook to the Water Works shared supply meters feeding Lakeview '
            '(Suberico, Guest Inn, and New Line / Police Container).'
        ),
        'calculation_notes': 'Shared production/distribution workbook sync using Water Works supply meters.',
        'assignments': [
            {
                'distribution_label': 'Suberico Meter',
                'meter_display_name': 'WWKS - 4" Suberico meter',
                'production_sheet': 'WATER WORKS',
                'production_label': '4" Suberico meter',
                'allocation_percentage': 100,
            },
            {
                'distribution_label': 'Guest Inn Meter',
                'meter_display_name': 'WWKS - Guest Inn meter',
                'production_sheet': 'WATER WORKS',
                'production_label': 'Guest Inn meter',
                'allocation_percentage': 100,
            },
            {
                'distribution_label': 'Police Container (Shared With Prodution)',
                'meter_display_name': 'WWKS - Police Container Meter',
                'production_sheet': 'WATER WORKS',
                'production_label': 'Police Container Meter (Faulty)',
                'allocation_percentage': 100,
            },
        ],
    },
    'MAIMAHIU': {
        'distribution_column': 'MAAIMAHIU',
        'description': 'Mapped from Distribution Zones workbook to the Mai Mahiu and Karima/NIP distribution stack.',
        'calculation_notes': 'Workbook sync using the listed Mai Mahiu area meters in the Maai Mahiu column.',
        'assignments': [
            {'distribution_label': 'NIP Gathima Meter (Ngujiri) (Meter over registering)', 'meter_display_name': 'NIP Gathima Meter', 'allocation_percentage': 100},
            {'distribution_label': 'Karima NIP 1', 'meter_display_name': 'Karima NIP 1', 'allocation_percentage': 100},
            {'distribution_label': 'Karima NIP 2', 'meter_display_name': 'Karima NIP 2', 'allocation_percentage': 100},
            {'distribution_label': 'Karima NIP 3', 'meter_display_name': 'Karima NIP 3', 'allocation_percentage': 100},
            {'distribution_label': 'Affordable Meter', 'meter_display_name': 'Affordable Meter', 'allocation_percentage': 100},
            {'distribution_label': 'Gathima Tank Outlet', 'meter_display_name': 'Gathima Tank Outlet', 'allocation_percentage': 100},
            {'distribution_label': 'ICD Dryport Meter (Faulty)', 'meter_display_name': 'ICD Dryport Meter', 'allocation_percentage': 100},
            {'distribution_label': 'Narok Meter', 'meter_display_name': 'Narok Meter', 'allocation_percentage': 100},
            {'distribution_label': 'Maai Mahiu BH1 Meter (shared with production)', 'meter_display_name': 'MAAI MAHIU BH1 METER', 'allocation_percentage': 100},
        ],
    },
    'KAYOLE': {
        'distribution_column': 'KAYOLE',
        'description': (
            'Mapped from Distribution Zones workbook to the Karati shared supply meters feeding Kayole '
            '(Water Works and Upper KWS).'
        ),
        'calculation_notes': 'Shared production/distribution workbook sync using Karati supply meters.',
        'assignments': [
            {
                'distribution_label': 'Water Works Meter',
                'meter_display_name': 'KARATI - WWKS METER',
                'production_sheet': 'KARATI',
                'production_label': 'WATER WORKS METER',
                'allocation_percentage': 100,
            },
            {
                'distribution_label': 'Upper KWS Meter',
                'meter_display_name': 'UPPER KWS - KAYOLE METER',
                'production_sheet': 'KARATI',
                'production_label': 'Upper KWS - KAYOLE',
                'allocation_percentage': 100,
            },
        ],
    },
    'IHINDU': {
        'distribution_column': 'IHINDU',
        'description': 'Mapped from Distribution Zones workbook to the two Ihindu shared production meters.',
        'calculation_notes': 'Shared production/distribution workbook sync using both Ihindu production meters.',
        'assignments': [
            {
                'distribution_label': 'Ihindu 1 Production Meter',
                'meter_display_name': 'IHINDU 1 BH1 METER',
                'production_sheet': 'IHINDU 1',
                'production_label': 'BH1 METER',
                'allocation_percentage': 100,
            },
            {
                'distribution_label': 'Ihindu 2 Production Meter',
                'meter_display_name': 'IHINDU 2 BH1 METER',
                'production_sheet': 'IHINDU 2',
                'production_label': 'BH1 METER',
                'allocation_percentage': 100,
            },
        ],
    },
    'SITE': {
        'distribution_column': 'DTI',
        'description': 'Mapped from Distribution Zones workbook to the Site and Services feeder meters under the DTI column.',
        'calculation_notes': 'Workbook sync using Site and Service, Upper Site, and Lower Site meters.',
        'assignments': [
            {'distribution_label': 'Site and Service Meter', 'meter_display_name': 'Site and Service Meter', 'allocation_percentage': 100},
            {'distribution_label': 'Upper Site Meter', 'meter_display_name': 'Upper Site Meter', 'allocation_percentage': 100},
            {'distribution_label': 'Lower site Meter', 'meter_display_name': 'Lower site Meter', 'allocation_percentage': 100},
        ],
    },
    'HOPEWELL': {
        'distribution_column': 'HOPEWELL',
        'description': 'Mapped from Distribution Zones workbook to the Karati Hopewell shared supply meter.',
        'calculation_notes': 'Shared production/distribution workbook sync using the Karati Hopewell supply meter.',
        'assignments': [
            {
                'distribution_label': 'Hopewell Meter',
                'meter_display_name': 'KARATI - HOPEWELL METER',
                'production_sheet': 'KARATI',
                'production_label': 'HOPEWELL METER',
                'allocation_percentage': 100,
            },
        ],
    },
    'GONDI': {
        'distribution_column': 'GONDI',
        'description': 'Mapped from Distribution Zones workbook to the shared Ngondi production/supply meter.',
        'calculation_notes': 'Shared production/distribution workbook sync using the Ngondi borehole meter.',
        'assignments': [
            {
                'distribution_label': 'Ngondi Meter',
                'meter_display_name': 'NGONDI BH1 METER',
                'production_sheet': 'NGONDI',
                'production_label': 'BH1 METER',
                'allocation_percentage': 100,
            },
        ],
    },
    'KINUNGI': {
        'distribution_column': 'KINUNGI',
        'description': 'Mapped from Distribution Zones workbook to the Kinungi production/distribution meter.',
        'calculation_notes': 'Workbook sync using the Kinungi BH production meter shown in the Kinungi column.',
        'assignments': [
            {
                'distribution_label': 'Kinungi BH/ Production Meter',
                'meter_display_name': 'KINUNGI BH1 METER',
                'allocation_percentage': 100,
            },
        ],
    },
    'NYONJORO': {
        'distribution_column': 'NYONJORO',
        'description': 'Mapped from Distribution Zones workbook to the shared Nyonjoro production/supply meter.',
        'calculation_notes': 'Shared production/distribution workbook sync using the Nyonjoro borehole meter.',
        'assignments': [
            {
                'distribution_label': 'Nyonjoro Meter',
                'meter_display_name': 'NYONJORO BH1 METER',
                'production_sheet': 'NYONJORO',
                'production_label': 'BH1 METER',
                'allocation_percentage': 100,
            },
        ],
    },
    'KAMERE': {
        'distribution_column': 'KAMERE',
        'description': 'Mapped from Distribution Zones workbook to the DCK shared supply meter feeding Kamere.',
        'calculation_notes': 'Shared production/distribution workbook sync using the DCK supply meter.',
        'assignments': [
            {
                'distribution_label': 'DCK BH1 meter',
                'meter_display_name': 'DCK Borehole 1 Meter',
                'production_sheet': 'DCK',
                'production_label': 'BH1 METER',
                'allocation_percentage': 100,
            },
        ],
    },
    'KABATI': {
        'distribution_column': 'DTI',
        'description': 'Mapped from Distribution Zones workbook to the Kabati inlet meter pair under the DTI column.',
        'calculation_notes': 'Workbook sync using the Kabati KAG and Makaburi meters already in the shared meter inventory.',
        'assignments': [
            {
                'distribution_label': 'Kabati, KAG meter',
                'meter_display_name': 'Kabati, KAG meter',
                'allocation_percentage': 100,
            },
            {
                'distribution_label': 'Kabati, Makaburi Meter',
                'meter_display_name': 'Kabati, Makaburi Meter',
                'allocation_percentage': 100,
            },
        ],
    },
    'LONGONOT': {
        'distribution_column': 'LONGONOT',
        'description': 'Mapped from Distribution Zones workbook to the Longonot main inlet meter.',
        'calculation_notes': 'Workbook sync using the canonical Longonot inlet meter only.',
        'assignments': [
            {
                'distribution_label': 'Longonot Main Meter',
                'meter_display_name': 'LONGONOT MAIN METER',
                'allocation_percentage': 100,
            },
        ],
    },
}


class Command(BaseCommand):
    help = (
        'Reconcile shared production/distribution supply meters from production sites.xlsx '
        'and Distribution Zones.xlsx into canonical zone assignment sets.'
    )

    def add_arguments(self, parser):
        parser.add_argument(
            '--production-file',
            default=None,
            help='Path to production sites.xlsx. Defaults to ../production sites.xlsx relative to manage.py.',
        )
        parser.add_argument(
            '--distribution-file',
            default=None,
            help='Path to Distribution Zones.xlsx. Defaults to ../Distribution Zones.xlsx relative to manage.py.',
        )
        parser.add_argument(
            '--meter-workbook',
            default=None,
            help='Path to WATER METERS UPDATED.xlsx used to restore imported display names after sync.',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Validate and report without writing to the database.',
        )

    @transaction.atomic
    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError('openpyxl is required to read the workbook files.') from exc

        production_file = options['production_file'] or os.path.normpath(
            os.path.join(os.path.dirname(os.path.abspath('manage.py')), '..', 'production sites.xlsx')
        )
        distribution_file = options['distribution_file'] or os.path.normpath(
            os.path.join(os.path.dirname(os.path.abspath('manage.py')), '..', 'Distribution Zones.xlsx')
        )
        meter_workbook = options['meter_workbook'] or os.path.normpath(
            os.path.join(os.path.dirname(os.path.abspath('manage.py')), '..', 'WATER METERS UPDATED.xlsx')
        )
        dry_run = options['dry_run']

        if not os.path.exists(production_file):
            raise CommandError(f'Production workbook not found: {production_file}')
        if not os.path.exists(distribution_file):
            raise CommandError(f'Distribution workbook not found: {distribution_file}')
        if not os.path.exists(meter_workbook):
            raise CommandError(f'Meter workbook not found: {meter_workbook}')

        production_workbook = openpyxl.load_workbook(production_file, data_only=True)
        distribution_workbook = openpyxl.load_workbook(distribution_file, data_only=True)
        imported_name_map = self._load_imported_meter_names(meter_workbook)

        distribution_labels = self._load_distribution_labels(distribution_workbook)
        production_labels = self._load_production_labels(production_workbook)

        updated_zones = 0
        created_assignments = 0

        for zone_code, plan in ZONE_SYNC_PLAN.items():
            zone = Zone.objects.filter(code=zone_code).first()
            if zone is None:
                self.stdout.write(self.style.WARNING(f'Skipping {zone_code}: zone does not exist in the current structure.'))
                continue

            self._validate_plan(zone_code, plan, distribution_labels, production_labels)

            if not dry_run:
                DistributionWaterMeterAssignment.objects.filter(zone=zone, is_active=True).update(is_active=False)
                DistributionWaterMeterAssignment.objects.filter(dma__zone=zone, is_active=True).update(is_active=False)

            config, _ = ZoneSupplyConfiguration.objects.get_or_create(zone=zone)
            if not dry_run:
                config.aggregation_method = 'CUSTOM_ASSIGNMENTS'
                config.primary_meter = None
                config.primary_water_meter = None
                config.infrastructure_description = plan['description']
                config.calculation_notes = plan['calculation_notes']
                config.save()
                config.component_dmas.clear()
                config.component_meters.clear()
                config.component_water_meters.clear()

            if not dry_run:
                zone.supply_aggregation_method = 'UNSET'
                zone.save(update_fields=['supply_aggregation_method', 'updated_at'])

            for item in plan['assignments']:
                water_meter = self._resolve_water_meter(item, imported_name_map)
                if water_meter is None:
                    raise CommandError(
                        f'Water meter not found for zone {zone_code} and label "{item["distribution_label"]}"'
                    )

                created = False
                if not dry_run:
                    assignment, created = DistributionWaterMeterAssignment.objects.update_or_create(
                        water_meter=water_meter,
                        zone=zone,
                        dma=None,
                        assignment_role='BULK_SUPPLY',
                        legacy_distribution_meter_id=None,
                        defaults={
                            'allocation_percentage': item['allocation_percentage'],
                            'is_active': True,
                            'start_date': None,
                            'end_date': None,
                            'notes': (
                                f'Workbook synced from Distribution Zones label "{item["distribution_label"]}". '
                                f'Canonical water meter {water_meter.meter_number}.'
                            ),
                        },
                    )
                    if not created:
                        assignment.is_active = True
                        assignment.save(update_fields=['is_active', 'updated_at'])
                    refresh_water_meter_display_name(water_meter)
                else:
                    existing = DistributionWaterMeterAssignment.objects.filter(
                        water_meter=water_meter,
                        zone=zone,
                        dma=None,
                        assignment_role='BULK_SUPPLY',
                        legacy_distribution_meter_id=None,
                    ).exists()
                    created = not existing
                if created:
                    created_assignments += 1

            updated_zones += 1

        restored_names = 0
        for meter_number, display_name in imported_name_map.items():
            meter = WaterMeter.objects.filter(meter_number=meter_number).first()
            if meter is None or meter.display_name == display_name:
                continue
            restored_names += 1
            if not dry_run:
                meter.display_name = display_name
                meter.save(update_fields=['display_name', 'updated_at'])

        unmatched_columns = self._find_unmapped_distribution_columns(distribution_labels)
        if unmatched_columns:
            self.stdout.write(self.style.WARNING(
                'Workbook columns not directly mapped by this command: ' + ', '.join(sorted(unmatched_columns))
            ))

        if dry_run:
            transaction.set_rollback(True)

        self.stdout.write(self.style.SUCCESS(
            f'Shared distribution metering sync complete. Updated {updated_zones} zones and created {created_assignments} active assignment rows. '
            f'Display names restored: {restored_names}.'
        ))

    def _load_distribution_labels(self, workbook):
        worksheet = workbook[workbook.sheetnames[0]]
        labels = {}
        for column in range(1, worksheet.max_column + 1):
            zone_title = worksheet.cell(row=1, column=column).value
            if not zone_title:
                continue
            zone_key = WORKBOOK_ZONE_ALIASES.get(normalize_text(zone_title), normalize_text(zone_title))
            labels[zone_key] = {
                str(row[0]).strip()
                for row in worksheet.iter_rows(min_row=2, min_col=column, max_col=column, values_only=True)
                if row[0]
            }
        return labels

    def _load_production_labels(self, workbook):
        labels = {}
        for worksheet in workbook.worksheets:
            sheet_labels = set()
            headers = {
                normalize_text(worksheet.cell(row=1, column=column).value): column - 1
                for column in range(1, worksheet.max_column + 1)
            }
            relevant_columns = [
                headers.get('PRODUCTIONWATERABSTRACTED'),
                headers.get('WATERSUPPLIED'),
            ]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                for index in relevant_columns:
                    if index is not None and index < len(row) and row[index]:
                        sheet_labels.add(str(row[index]).strip())
            labels[worksheet.title] = sheet_labels
        return labels

    def _load_imported_meter_names(self, workbook_path):
        import openpyxl

        workbook = openpyxl.load_workbook(workbook_path, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        mapping = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            raw_value = row[0] if row else None
            raw_name = str(raw_value).strip() if raw_value is not None else ''
            if not raw_name:
                continue
            clean_name, _ = split_meter_name_and_notes(raw_name)
            mapping[slugify_meter_name(raw_name)] = clean_name
            mapping[slugify_meter_name(clean_name)] = clean_name
        return mapping

    def _resolve_water_meter(self, item, imported_name_map):
        direct_name = item.get('meter_display_name')
        if direct_name:
            meter = WaterMeter.objects.filter(display_name=direct_name).first()
            if meter:
                return meter

        direct_number = item.get('meter_number')
        if direct_number:
            meter = WaterMeter.objects.filter(meter_number=direct_number).first()
            if meter:
                return meter
            cleaned_from_number = imported_name_map.get(direct_number)
            if cleaned_from_number:
                meter = WaterMeter.objects.filter(display_name=cleaned_from_number).first()
                if meter:
                    return meter

        labels = [
            item.get('distribution_label'),
            item.get('production_label'),
            direct_name,
        ]
        labels = [label for label in labels if label]

        candidates = []
        for label in labels:
            clean_label, _ = split_meter_name_and_notes(label)
            variants = {normalize_text(label), normalize_text(clean_label)}
            variants = {variant for variant in variants if variant}

            for meter in WaterMeter.objects.order_by('display_name'):
                meter_key = normalize_text(meter.display_name)
                if any(
                    meter_key == variant or meter_key.endswith(variant) or variant.endswith(meter_key) or variant in meter_key
                    for variant in variants
                ):
                    candidates.append(meter)

        unique = []
        seen_ids = set()
        for meter in candidates:
            if meter.id in seen_ids:
                continue
            seen_ids.add(meter.id)
            unique.append(meter)

        if len(unique) == 1:
            return unique[0]
        if len(unique) > 1:
            raise CommandError(
                f'Ambiguous distribution meter match for "{item.get("distribution_label")}". Candidates: '
                + ', '.join(meter.display_name for meter in unique)
            )
        return None

    def _validate_plan(self, zone_code, plan, distribution_labels, production_labels):
        zone_distribution_labels = distribution_labels.get(plan.get('distribution_column', zone_code), set())
        if not zone_distribution_labels:
            raise CommandError(
                f'Distribution workbook column for zone {plan.get("distribution_column", zone_code)} was not found.'
            )

        for item in plan['assignments']:
            distribution_column = item.get('distribution_column', zone_code)
            labels_for_item = distribution_labels.get(distribution_column, zone_distribution_labels)
            if not self._label_exists(item['distribution_label'], labels_for_item):
                raise CommandError(
                    f'Distribution label "{item["distribution_label"]}" not found in workbook column for zone {distribution_column}.'
                )

            sheet_name = item.get('production_sheet')
            production_label = item.get('production_label')
            if sheet_name and production_label:
                if sheet_name not in production_labels:
                    raise CommandError(f'Production workbook sheet "{sheet_name}" not found.')
                if not self._label_exists(production_label, production_labels[sheet_name]):
                    raise CommandError(
                        f'Production label "{production_label}" not found on workbook sheet "{sheet_name}".'
                    )

    def _label_exists(self, expected_label, available_labels):
        expected_clean, _ = split_meter_name_and_notes(expected_label)
        expected_variants = {
            normalize_text(expected_label),
            normalize_text(expected_clean),
        }
        expected_variants = {variant for variant in expected_variants if variant}

        for available_label in available_labels:
            available_clean, _ = split_meter_name_and_notes(available_label)
            available_variants = {
                normalize_text(available_label),
                normalize_text(available_clean),
            }
            available_variants = {variant for variant in available_variants if variant}
            for expected_variant in expected_variants:
                for available_variant in available_variants:
                    if (
                        expected_variant == available_variant
                        or expected_variant in available_variant
                        or available_variant in expected_variant
                    ):
                        return True
        return False

    def _find_unmapped_distribution_columns(self, distribution_labels):
        mapped = {plan['distribution_column'] for plan in ZONE_SYNC_PLAN.values()}
        unmapped = []
        for workbook_key in distribution_labels:
            if workbook_key not in ZONE_SYNC_PLAN:
                unmapped.append(workbook_key)
        return unmapped
