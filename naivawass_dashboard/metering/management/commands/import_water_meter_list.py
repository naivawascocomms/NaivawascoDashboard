from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from metering.models import WaterMeter
from metering.workbook_names import (
    build_workbook_annotation,
    infer_meter_operational_status,
    merge_notes,
    slugify_meter_name,
    split_meter_name_and_notes,
)


class Command(BaseCommand):
    help = 'Import a one-column water meter list workbook into the shared WaterMeter table.'

    def add_arguments(self, parser):
        parser.add_argument(
            'workbook',
            nargs='?',
            default=None,
            help='Path to the .xlsx workbook. Defaults to ../WATER METERS UPDATED.xlsx relative to manage.py.',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Parse the workbook and report what would be imported without writing to the database.',
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError('openpyxl is required to import the workbook.') from exc

        workbook_path = options['workbook']
        if workbook_path is None:
            workbook_path = Path.cwd().parent / 'WATER METERS UPDATED.xlsx'
        workbook_path = Path(workbook_path).resolve()

        if not workbook_path.exists():
            raise CommandError(f'Workbook not found: {workbook_path}')

        dry_run = options['dry_run']
        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN - no database changes will be made.'))

        workbook = openpyxl.load_workbook(workbook_path, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]

        meter_entries = []
        seen_names = set()
        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            raw_value = row[0] if row else None
            raw_name = str(raw_value).strip() if raw_value is not None else ''
            if not raw_name:
                continue
            if raw_name in seen_names:
                raise CommandError(f'Duplicate meter name "{raw_name}" found in workbook row {row_index}.')
            seen_names.add(raw_name)
            clean_name, workbook_notes = split_meter_name_and_notes(raw_name)
            meter_entries.append({
                'raw_name': raw_name,
                'clean_name': clean_name,
                'workbook_notes': workbook_notes,
            })

        if not meter_entries:
            raise CommandError('No meter names found in workbook.')

        installation_date = timezone.localdate()
        import_note = (
            f'Imported from {workbook_path.name} on {installation_date.isoformat()}. '
            'Installation date was not provided in the workbook; import date stored as placeholder.'
        )

        created = 0
        updated = 0

        with transaction.atomic():
            for entry in meter_entries:
                raw_name = entry['raw_name']
                clean_name = entry['clean_name']
                workbook_annotation = build_workbook_annotation(entry['workbook_notes'])
                operational_status = infer_meter_operational_status(raw_name, workbook_annotation)
                meter_number = slugify_meter_name(clean_name)
                original_meter_number = meter_number
                suffix = 2
                while WaterMeter.objects.exclude(display_name=clean_name).filter(meter_number=meter_number).exists():
                    suffix_text = f'-{suffix}'
                    meter_number = f'{original_meter_number[:100 - len(suffix_text)]}{suffix_text}'
                    suffix += 1

                raw_meter_number = slugify_meter_name(raw_name)
                meter = WaterMeter.objects.filter(
                    meter_number__in={meter_number, raw_meter_number}
                ).first()
                if meter is None:
                    meter = WaterMeter.objects.filter(display_name__in={clean_name, raw_name}).first()
                if meter:
                    changed = False
                    if not meter.meter_number:
                        meter.meter_number = meter_number
                        changed = True
                    if meter.display_name != clean_name:
                        meter.display_name = clean_name
                        changed = True
                    if not meter.installation_date:
                        meter.installation_date = installation_date
                        changed = True
                    desired_notes = merge_notes(import_note, workbook_annotation)
                    if desired_notes and desired_notes not in (meter.notes or ''):
                        meter.notes = merge_notes(meter.notes, desired_notes)
                        changed = True
                    if meter.operational_status != operational_status:
                        meter.operational_status = operational_status
                        changed = True
                    if workbook_annotation and workbook_annotation not in (meter.operational_status_notes or ''):
                        meter.operational_status_notes = merge_notes(
                            meter.operational_status_notes,
                            workbook_annotation,
                        )
                        changed = True
                    if not meter.is_active:
                        meter.is_active = True
                        changed = True
                    if changed and not dry_run:
                        meter.save()
                        updated += 1
                    continue

                if dry_run:
                    created += 1
                    continue

                WaterMeter.objects.create(
                    meter_number=meter_number,
                    display_name=clean_name,
                    installation_date=installation_date,
                    initial_reading=0,
                    notes=merge_notes(import_note, workbook_annotation),
                    operational_status=operational_status,
                    operational_status_notes=workbook_annotation,
                    is_active=True,
                )
                created += 1

            if dry_run:
                transaction.set_rollback(True)

        self.stdout.write(
            self.style.SUCCESS(
                f'Processed {len(meter_entries)} workbook entries. Created: {created}. Updated: {updated}.'
            )
        )
